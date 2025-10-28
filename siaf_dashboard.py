# -*- coding: utf-8 -*-

import base64
import io
import re
from contextlib import contextmanager
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Tuple

import altair as alt
import numpy as np
import pandas as pd
import streamlit as st

alt.renderers.set_embed_options(
    actions=False,
    formatLocale={
        "decimal": ".",
        "thousands": ",",
        "grouping": [3],
        "currency": ["S/. ", ""],
    },
)

EXCEL_SOURCE_DIR = Path(__file__).parent / "data" / "siaf"
EXCEL_SOURCE_DIR.mkdir(parents=True, exist_ok=True)


def _list_excel_candidates(folder: Path) -> List[Path]:
    return sorted(
        [p for p in folder.glob("*.xlsx") if p.is_file()],
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )


EXCEL_CANDIDATES = _list_excel_candidates(EXCEL_SOURCE_DIR)
LATEST_EXCEL = EXCEL_CANDIDATES[0] if EXCEL_CANDIDATES else None

PRIMARY_COLOR = "#ff296d"
SECONDARY_COLOR = "#0b0f2f"
ACCENT_COLOR = "#00eaff"
GLOW_COLOR = "#b388ff"

ARCADE_COLOR_RANGE = [
    "#00eaff",
    "#ff296d",
    "#ffe45c",
    "#19f5aa",
    "#b388ff",
    "#ff9d00",
]

try:
    import xlsxwriter  # type: ignore
except ModuleNotFoundError:
    XLSXWRITER_AVAILABLE = False
else:
    XLSXWRITER_AVAILABLE = True

try:
    import openpyxl  # type: ignore # noqa: F401
except ModuleNotFoundError:
    OPENPYXL_AVAILABLE = False
else:
    OPENPYXL_AVAILABLE = True

EXCEL_CURRENCY_FORMAT = '"S/." #,##0.00'
EXCEL_PERCENT_FORMAT = "0.00%"

LOGO_BASE64 = (
    "iVBORw0KGgoAAAANSUhEUgAAAMgAAADICAYAAACtWK6eAAAIJklEQVR4nO3cO4uUVxjA8We8bLISIiRxC1Mu2AleChEE11L8ALZi6WewtrK1URsr/QZWgo2VjSBWYiw0"
    "G7AIIUiCqzApNq/OrjvH93Iuz+X/q0KEdfac85/nzLizs/l8LgD2tq/1AwA0IxAggUCABAIBEggESCAQIIFAgAQCARIIBEg40PoBRPJqNsv2Ywvr8/ks19fCcjN+1CSv"
    "nBGMRTz5EMgEGmLoi2jGIZCBLEWxDLH0RyDf4CGIbyGY5QhkiQhh7EYoXyOQBRGjWIZYthGIEEZK9FDCBkIUw0WMJVwghDFdpFDCBEIY+UUIxX0ghFGe51Bc/7AicdTh"
    "eZ1dThDPG6adt2niboIQR1ve1t/NBPG2MR54mCYuJghx6ORhX0xPEA8bEIXVaWJ2ghCHLVb3y2QgVhc7Oov7ZuqKZXGBsTcrVy4zE4Q4fLGynyYCsbKYGMbCvqoPxMIi"
    "Yjzt+6s6EO2Lhzw077PaQDQvGvLTut8qA9G6WChL476rC0TjIqEebfuvKhBti4M2NJ0DNYFoWhS0p+U8qAhEy2JAFw3nonkgGhYBerU+H00Daf3Nw4aW56RZIMSBIVqd"
    "l+ZXLECzJoEwPTBGi3NTPRDiwBS1z0/VQIgDOdQ8R7wGARKqBcL0QE61zlOVz6QTx07rE9b81czER7mrKf3Z9gMlvzimxdD36xFNOcUnSMTpkTuKISLGUnKKFA0kUhwt"
    "o1gmUiylIuGKNZHGMDrdY4sUSm7FJoj36aE5jGW8h1JiihQJxHMcFsPYzXMouSPhitWThzA6XL36y/4PhR6nh6c4Fnn8vnKfPyZIgscDtBvTJC3rBPE0PSLEscjT95vz"
    "HPLDinvwdFiGiPp9p2R7F8vD9OCAfOHhypXjHS0myP+IYyfWYxuBCIdhGdYlUyCWr1ccgjTL65PjXIaeIJY3v6bI6zQ5EKvTI/Kmj2F1vaaez9ATBPiWkIFYfTZsLeK6"
    "TQrE4vUq4ibnZHH9ppzTUBPE4uZqFGkdRwdicXogrrHnNcwEifSsV0OU9QwRSJTNrC3CuoYIBBhrVCCWXn9EeJZrydL6jjm3TBAgwXUglp7dLPO8zoMDsXS9AnYben7d"
    "ThDPz2oaeV1vt4EAObgMJNez2W+HDsnmxoZsXrggv587Jx+ePhURkb/v3JG3p07J5vnz8selS/LpzZssf591HqcIvxcrYbayIkcfPxYRka3nz+Xd1avy840b8v7+ffn1"
    "yROZra7KPw8fyrsrV+Too0dtHyyKGDRBIr9AXzl+XD69fi1/3bwpP924IbPVVREROXTxohxcX5f5x4+NHyH6GnKO3V2xSo35fx89kpUTJ2TrxQv57uTJHX925PZtmR08"
    "WOTvtcbbNYsrVsJ8a0s2NzZE5nPZd/iwrN29K2/Pnm39sFARgSQsvgbprBw7Jh+ePZPvz5zZ/h/zuby7ckXW7t2r/wBRnKsrVo3x/uO1a/Ln9esy//BBRETeP3jw+b+x"
    "zdM1iwky0A+XL8vHly/l7enTsv/IEdm/tia/3LrV+mGhkEG/m1f7u1ienrms0/67ffv+3t7eVyztcQBD9D3Prl6DALm5CYTrlS5e9sNNIEAJBAIkEAiQQCBAAoEACQQC"
    "JBAIkEAgQAKBAAkEAiQQCJBAIEACgQAJBAIkuAlE+yfYovGyH24CAUroHUjfz/ACFmT/TDoQEYEACa4C8fLC0DpP++AqECA3AgES3AXiabxb5G39BwXCW73wYMg5djdB"
    "gJxcBuJtzFvhcd1dBgLk4jYQj89mmnld78GB8EIdlg09v24niIjfZzVtPK+z60CAqUYFYuma5fnZTQNL6zvm3DJBgIQQgVh6lrMkwrqGCEQkxmbWFGU9Rwdi6XUIMPa8"
    "hpkgInGe9UqLtI6TArE4RSJtbgkW12/KOQ01QToWN1mDiOsWMhCgr8mBWLxmicR8NpzC6npNPZ+hJ4jVTa8t8jplCcTqFBGJvfl9WF6fHOcy9ATpWD4EJbEuBPIZh2En"
    "1mPbbD6fZ/tir2azfF+sofWMa2KNlzByXfuZIHvwckiGivp9p2QNxPKL9d2iHRZP32/Oc3gg1xfyqDs0nq9cnsIoIfsVy9MU6Xg9RB6/r9znjwnSk6dp4jGMUrK+i7XI"
    "yztay1gMxXsYJW4vxQIR8R+JiI1QvIchUu5qzxVrIs1XrwhhlFZ0gojEmCK7tYwlYhQl3xgqHohIzEg6NWKJGEWn9LumXLEK2+vwTokmcgwtVJkgIrGnCMqo8W9u1X4W"
    "y+M/IKKdWueJH1YEEqoGwhRBDjXPUfUJQiSYovb5aXLFIhKM0eLc8BoESGgWCFMEQ7Q6L00nCJGgj5bnpPkVi0iQ0vp8NA9EpP0iQCcN50JFICI6FgN6aDkPagIR0bMo"
    "aEvTOVAViIiuxUF92vZfXSAi+hYJdWjcd5WBiOhcLJSjdb/VBiKid9GQl+Z9Vh2IiO7Fw3Ta91d9ICL6FxHjWNhXE4GI2FhM9GdlP6t9Jj0nPt9ul5UwOmYmyCJri4xt"
    "FvfNZCAiNhc7Mqv7ZfKKtRtXLr2shtExO0EWWd8Erzzsi4sJsohp0p6HMDouJsgiT5tjkbf1dzdBFjFN6vEWRsfdBFnkddO08bzOrifIIqZJfp7D6IQJpEMo00UIoxMu"
    "kA6hDBcpjE7YQBYRy3IRo1hEIAsI5YvoYXQIZImIsRDF1wjkGyKEQhjLEchAHoIhiP4IZAJLsRDFOASSmYZoiCEfAqkoZzxEUAeBAAmuf1gRmIpAgAQCARIIBEggECCB"
    "QIAEAgESCARI+A/Mh09abbhiGAAAAABJRU5ErkJggg=="
)
LOGO_IMAGE = base64.b64decode(LOGO_BASE64)

APP_CSS = f"""
<style>
@import url('https://fonts.googleapis.com/css2?family=Orbitron:wght@400;600;700&family=Rajdhani:wght@400;600;700&display=swap');

:root {{
    --primary-color: {PRIMARY_COLOR};
    --accent-color: {ACCENT_COLOR};
    --secondary-color: {SECONDARY_COLOR};
    --surface-color: rgba(16, 21, 46, 0.92);
    --panel-border: rgba(0, 234, 255, 0.35);
    --glow-color: {GLOW_COLOR};
}}

html, body, [data-testid="stAppViewContainer"] {{
    background: radial-gradient(circle at 20% 20%, rgba(0, 234, 255, 0.18), rgba(11, 15, 47, 0.95) 55%) fixed;
    color: #f1f5ff;
    font-family: 'Rajdhani', sans-serif;
}}

[data-testid="stSidebar"] {{
    background: linear-gradient(180deg, rgba(19, 32, 67, 0.95), rgba(11, 15, 47, 0.85));
    border-right: 1px solid rgba(0, 234, 255, 0.25);
}}

[data-testid="stSidebar"] * {{
    color: #e3f6ff !important;
}}

.app-title {{
    font-family: 'Orbitron', sans-serif;
    font-size: 2.8rem;
    font-weight: 700;
    margin-bottom: 0.1rem;
    color: var(--accent-color);
    text-shadow: 0 0 18px rgba(0, 234, 255, 0.65);
}}

.app-subtitle {{
    color: #f06292;
    font-size: 1.15rem;
    margin-top: 0;
    margin-bottom: 0.6rem;
    letter-spacing: 0.12rem;
}}

.app-description {{
    color: #d0dcff;
    font-size: 1.02rem;
    line-height: 1.6rem;
    background: rgba(15, 20, 45, 0.65);
    padding: 0.85rem 1.2rem;
    border-radius: 12px;
    border: 1px solid rgba(0, 234, 255, 0.15);
    box-shadow: inset 0 0 14px rgba(255, 41, 109, 0.18);
}}

.stTabs [data-baseweb="tab"] {{
    color: rgba(227, 246, 255, 0.85);
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: 0.08rem;
    font-family: 'Orbitron', sans-serif;
}}

.stTabs [data-baseweb="tab"]:hover {{
    color: var(--accent-color);
}}

.stTabs [data-baseweb="tab"][aria-selected="true"] {{
    color: var(--primary-color);
    border-bottom: 4px solid var(--accent-color);
    text-shadow: 0 0 8px rgba(0, 234, 255, 0.55);
}}

[data-testid="stMetricValue"] {{
    color: #ffe45c;
    font-family: 'Orbitron', sans-serif;
    text-shadow: 0 0 15px rgba(255, 228, 92, 0.55);
}}

[data-testid="stMetricLabel"] {{
    color: rgba(227, 246, 255, 0.75);
    font-weight: 600;
    letter-spacing: 0.06rem;
}}

.stRadio > div {{
    background-color: rgba(0, 234, 255, 0.08);
    border-radius: 999px;
    padding: 0.4rem 0.8rem;
    border: 1px solid rgba(0, 234, 255, 0.25);
}}

.stRadio [data-baseweb="radio"] label span {{
    font-weight: 600;
    color: rgba(227, 246, 255, 0.85);
}}

.stRadio [data-baseweb="radio"] input:checked + span {{
    color: var(--accent-color);
    text-shadow: 0 0 10px rgba(0, 234, 255, 0.65);
}}

.stButton>button, .stDownloadButton>button {{
    background: linear-gradient(135deg, var(--accent-color), var(--primary-color));
    color: #ffffff;
    font-weight: 700;
    border: 1px solid rgba(255, 255, 255, 0.25);
    box-shadow: 0 12px 28px rgba(0, 234, 255, 0.25);
    border-radius: 999px;
    text-transform: uppercase;
    letter-spacing: 0.08rem;
}}

.stButton>button:hover, .stDownloadButton>button:hover {{
    box-shadow: 0 18px 42px rgba(255, 41, 109, 0.35);
}}

.neon-card {{
    background: linear-gradient(145deg, rgba(11, 16, 38, 0.96), rgba(21, 37, 71, 0.92));
    border: 1px solid var(--panel-border);
    border-radius: 18px;
    padding: 1.6rem 1.8rem;
    margin-bottom: 1.4rem;
    position: relative;
    overflow: hidden;
    box-shadow: 0 18px 32px rgba(0, 0, 0, 0.35), inset 0 0 22px rgba(0, 234, 255, 0.08);
}}

.neon-card:before {{
    content: "";
    position: absolute;
    inset: -80px;
    background: conic-gradient(from 0deg, rgba(0, 234, 255, 0.35), transparent 45%, rgba(255, 41, 109, 0.35), transparent 75%);
    filter: blur(60px);
    opacity: 0.5;
    animation: pulseGlow 12s linear infinite;
}}

.neon-card:after {{
    content: "";
    position: absolute;
    inset: 2px;
    border-radius: 16px;
    background: rgba(5, 8, 22, 0.88);
    box-shadow: inset 0 0 18px rgba(0, 234, 255, 0.08);
}}

.neon-card > * {{
    position: relative;
    z-index: 2;
}}

.panel-header {{
    display: flex;
    align-items: center;
    gap: 0.75rem;
    margin-bottom: 1.1rem;
}}

.panel-icon {{
    width: 48px;
    height: 48px;
    border-radius: 50%;
    display: grid;
    place-items: center;
    font-size: 1.6rem;
    background: rgba(0, 234, 255, 0.15);
    border: 1px solid rgba(0, 234, 255, 0.45);
    box-shadow: 0 0 18px rgba(0, 234, 255, 0.35);
}}

.panel-title {{
    font-family: 'Orbitron', sans-serif;
    font-size: 1.55rem;
    letter-spacing: 0.08rem;
    text-transform: uppercase;
}}

.panel-subtitle {{
    color: rgba(227, 246, 255, 0.7);
    margin-top: -0.8rem;
    margin-bottom: 1.1rem;
    font-size: 0.95rem;
    letter-spacing: 0.05rem;
}}

.panel-body {{
    position: relative;
    z-index: 2;
}}

.analysis-bubble {{
    margin-top: 1.1rem;
    padding: 0.9rem 1.1rem;
    border-radius: 14px;
    background: rgba(255, 41, 109, 0.08);
    border: 1px solid rgba(255, 41, 109, 0.45);
    box-shadow: 0 0 18px rgba(255, 41, 109, 0.25);
    font-weight: 600;
    display: flex;
    gap: 0.75rem;
    align-items: flex-start;
    color: #ffdce6;
}}

.analysis-icon {{
    font-size: 1.2rem;
    margin-top: 0.2rem;
}}

.neon-subheader {{
    font-family: 'Orbitron', sans-serif;
    text-transform: uppercase;
    letter-spacing: 0.12rem;
    color: var(--accent-color);
    margin-top: 1.4rem;
    margin-bottom: 0.6rem;
    text-shadow: 0 0 12px rgba(0, 234, 255, 0.55);
}}

.stSlider > div {{
    padding: 0.4rem 0.7rem;
    background: rgba(0, 234, 255, 0.06);
    border-radius: 14px;
    border: 1px solid rgba(0, 234, 255, 0.18);
}}

.stSlider [data-baseweb="slider"] > div > div {{
    background: linear-gradient(90deg, rgba(0, 234, 255, 0.75), rgba(255, 41, 109, 0.75));
}}

.stSelectbox label, .stMultiselect label {{
    text-transform: uppercase;
    font-family: 'Orbitron', sans-serif;
    letter-spacing: 0.08rem;
}}

.stDataFrame {{
    filter: drop-shadow(0 12px 24px rgba(0, 0, 0, 0.45));
}}

.stDataFrame table {{
    border-radius: 12px;
    overflow: hidden;
}}

.stDataFrame [role="gridcell"], .stDataFrame [role="columnheader"] {{
    color: #f1f5ff !important;
    background-color: rgba(9, 14, 29, 0.85) !important;
    border-color: rgba(0, 234, 255, 0.12) !important;
}}

.stDataFrame [role="columnheader"] {{
    font-family: 'Orbitron', sans-serif;
    letter-spacing: 0.05rem;
}}

@keyframes pulseGlow {{
    0% {{ transform: rotate(0deg); opacity: 0.4; }}
    50% {{ transform: rotate(180deg); opacity: 0.8; }}
    100% {{ transform: rotate(360deg); opacity: 0.4; }}
}}
</style>"""

# =========================
# Configuraci√≥n de la app
# =========================
st.set_page_config(page_title="SIAF Dashboard - Peru Compras", layout="wide")
st.markdown(APP_CSS, unsafe_allow_html=True)


@contextmanager
def neon_panel(title: str, icon: str = "üïπÔ∏è", subtitle: str | None = None):
    st.markdown("<div class='neon-card'>", unsafe_allow_html=True)
    subtitle_html = f"<div class='panel-subtitle'>{subtitle}</div>" if subtitle else ""
    st.markdown(
        "<div class='panel-header'>"
        f"<span class='panel-icon'>{icon}</span>"
        f"<div><div class='panel-title'>{title}</div>{subtitle_html}</div></div>"
        "<div class='panel-body'>",
        unsafe_allow_html=True,
    )
    try:
        yield
    finally:
        st.markdown("</div></div>", unsafe_allow_html=True)


def neon_subheader(text: str, icon: str = "‚ú®") -> None:
    st.markdown(
        f"<h3 class='neon-subheader'><span>{icon}</span> {text}</h3>",
        unsafe_allow_html=True,
    )


def render_analysis(text: str, icon: str = "üß†") -> None:
    st.markdown(
        f"<div class='analysis-bubble'><span class='analysis-icon'>{icon}</span><span>{text}</span></div>",
        unsafe_allow_html=True,
    )


def style_arcade_chart(chart: alt.Chart) -> alt.Chart:
    return (
        chart.configure_view(
            strokeOpacity=0,
            fill="#0b1126",
        )
        .configure_axis(
            labelColor="#e3f6ff",
            titleColor="#78f1ff",
            gridColor="rgba(120, 241, 255, 0.12)",
            domainColor="rgba(120, 241, 255, 0.28)",
        )
        .configure_legend(
            labelColor="#e3f6ff",
            titleColor="#78f1ff",
            orient="bottom",
            direction="horizontal",
            symbolType="stroke",
            symbolSize=150,
        )
        .configure_title(
            font="Orbitron",
            color="#e3f6ff",
            fontWeight="bold",
        )
    )

header_col_logo, header_col_text = st.columns([1, 4])
with header_col_logo:
    st.image(LOGO_IMAGE, width=120)
with header_col_text:
    st.markdown("<h1 class='app-title'>SIAF Dashboard - Per√∫ Compras</h1>", unsafe_allow_html=True)
    st.markdown(
        "<p class='app-subtitle'>Seguimiento diario del avance de ejecuci√≥n presupuestal</p>",
        unsafe_allow_html=True,
    )

st.markdown(
    "<p class='app-description'>El dashboard toma autom√°ticamente el <strong>Excel SIAF</strong> m√°s reciente de la carpeta "
    "<code>data/siaf</code> para analizar <strong>PIA, PIM, Certificado, No certificado, Comprometido, Devengado y % de avance</strong>. "
    "La aplicaci√≥n asegura la lectura completa hasta CI, construye clasificadores jer√°rquicos estandarizados y ofrece vistas din√°micas con descargas.</p>",
    unsafe_allow_html=True,
)

# =========================
# Sidebar / par√°metros
# =========================
selected_excel_path = LATEST_EXCEL

with st.sidebar:
    st.image(LOGO_IMAGE, width=140)
    st.markdown("<h3 style='color: var(--primary-color); margin-top: 0.5rem;'>Panel de control</h3>", unsafe_allow_html=True)
    st.header("Origen de datos")
    st.caption(
        "Coloca los archivos <code>.xlsx</code> en <code>data/siaf</code>. El dashboard usa el m√°s reciente autom√°ticamente."
    )
    if not EXCEL_CANDIDATES:
        st.error("No se encontraron archivos .xlsx en data/siaf. A√±ade uno y vuelve a actualizar.")
    else:
        label_to_path = {}
        option_labels = []
        for path in EXCEL_CANDIDATES:
            updated = datetime.fromtimestamp(path.stat().st_mtime).strftime("%d/%m/%Y %H:%M")
            label = f"{path.name} ¬∑ {updated}"
            label_to_path[label] = path
            option_labels.append(label)
        selected_label = st.selectbox(
            "Selecciona el archivo SIAF",
            options=option_labels,
            index=0,
            help="Los archivos est√°n ordenados del m√°s reciente al m√°s antiguo.",
        )
        selected_excel_path = label_to_path[selected_label]
        st.success(f"Usando: {selected_excel_path.name}")
    st.markdown("---")
    st.header("Par√°metros de lectura")
    usecols = st.text_input(
        "Rango de columnas (Excel)",
        "A:DV",
        help="Lectura fija para asegurar columnas CI‚ÄìEC y programaci√≥n mensual",
        disabled=True,
    )
    sheet_name = st.text_input("Nombre de hoja (opcional)", "", help="D√©jalo vac√≠o para autodetecci√≥n.")
    header_row_excel = st.number_input("Fila de encabezados (Excel, 1=primera)", min_value=1, value=4)
    detect_header = st.checkbox("Autodetectar encabezado", value=True)
    st.markdown("---")
    st.header("Reglas CI‚ÄìEC")
    current_month = st.number_input("Mes actual (1-12)", min_value=1, max_value=12, value=9)
    riesgo_umbral = st.number_input("Umbral de avance m√≠nimo (%)", min_value=0, max_value=100, value=60)
    meta_avance = st.number_input("Meta de avance al cierre (%)", min_value=0, max_value=100, value=95)
    st.caption("Se marca riesgo_devolucion si Avance% < Umbral.")

if selected_excel_path is None:
    st.error("No hay archivos disponibles en data/siaf. A√±ade un Excel y vuelve a ejecutar el dashboard.")
    st.stop()

# Mapeo de c√≥digos de sec_func a nombres
SEC_FUNC_MAP = {
    1: "PI_2",
    2: "DCEME",
    3: "DE",
    4: "PI_1",
    5: "OPP",
    6: "JEF",
    7: "GG",
    8: "OAUGD",
    9: "OTI",
    10: "OA",
    11: "OC",
    12: "OAJ",
    13: "RRHH",
    14: "OCI",
    15: "DCEME15",
    16: "DETN16",
    21: "DETN21",
    22: "DETN22",
}
SEC_FUNC_MAP.update({str(k): v for k, v in SEC_FUNC_MAP.items()})

_sec_func_pattern = re.compile(r"^\s*0*(\d+)")


def map_sec_func(value):
    """Normaliza y reemplaza los c√≥digos *sec_func* por sus √°reas."""
    if pd.isna(value):
        return value

    if isinstance(value, (int, np.integer)):
        key = int(value)
        return SEC_FUNC_MAP.get(key, SEC_FUNC_MAP.get(str(key), value))

    if isinstance(value, float) and value.is_integer():
        key = int(value)
        return SEC_FUNC_MAP.get(key, SEC_FUNC_MAP.get(str(key), value))

    text = str(value).strip()
    if not text:
        return value

    match = _sec_func_pattern.match(text)
    if match:
        key_str = match.group(1)
        key_int = int(key_str)
        mapped = SEC_FUNC_MAP.get(key_int, SEC_FUNC_MAP.get(key_str))
        if mapped is not None:
            return mapped

    return SEC_FUNC_MAP.get(text, value)


AMOUNT_KEYWORDS = (
    "mto",
    "devengado",
    "saldo",
    "pia",
    "pim",
    "certificado",
    "compro",
    "monto",
    "actual",
    "real",
    "necesario",
    "estimado",
    "proyeccion",
    "programado",
)
EXCLUDE_ROUND_COLS = {"mes", "rank_acum", "rank_mes", "n"}
Z_SCORE_95 = 1.96

FINANCIAL_RENAME_MAP = {
    "mto_pia": "PIA",
    "mto_pim": "PIM",
    "mto_certificado": "CERTIFICADO",
    "mto_compro_anual": "COMPROMETIDO",
    "devengado": "DEVENGADO",
    "AVANCE DE EJECUCI√ìN ACUMULADO": "DEVENGADO",
    "devengado_mes": "DEVENGADO MES",
    "programado_mes": "PROGRAMADO MES",
    "avance_%": "AVANCE",
    "avance_acum_%": "AVANCE",
    "% AVANCE DEV /PIM": "AVANCE",
    "Avance%": "AVANCE",
    "avance_programado_%": "AVANCE MES",
    "% AVANCE DEV MES/PROG": "AVANCE MES",
    "AvanceProgramado%": "AVANCE MES",
    "avance_mes_%": "AVANCE MES (PIM)",
    "no_certificado": "NO CERTIFICADO",
    "saldo_pim": "NO CERTIFICADO",
}
FINANCIAL_ORDER = ("PIM", "CERTIFICADO", "COMPROMETIDO", "DEVENGADO", "AVANCE")
FINANCIAL_MONTHLY_ORDER = ("DEVENGADO MES", "PROGRAMADO MES", "AVANCE MES")
FINANCIAL_FINAL_COLUMN = "NO CERTIFICADO"
PERCENT_DISPLAY_COLUMNS = {"AVANCE", "AVANCE MES"}


def standardize_financial_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Rename and reorder budget execution columns following the requested layout."""

    if df is None or df.empty:
        return df

    df = df.copy()
    rename_map = {
        source: target
        for source, target in FINANCIAL_RENAME_MAP.items()
        if source in df.columns and source != target
    }
    if rename_map:
        df = df.rename(columns=rename_map)

    columns = list(df.columns)
    target_set = set(FINANCIAL_ORDER) | set(FINANCIAL_MONTHLY_ORDER) | {FINANCIAL_FINAL_COLUMN}
    leading_columns = [c for c in columns if c not in target_set]
    ordered = leading_columns
    ordered.extend([c for c in FINANCIAL_ORDER if c in df.columns])
    ordered.extend([c for c in FINANCIAL_MONTHLY_ORDER if c in df.columns])
    if FINANCIAL_FINAL_COLUMN in df.columns:
        ordered.append(FINANCIAL_FINAL_COLUMN)

    # Ensure all columns are present in the final layout without duplication.
    seen = set()
    final_order = []
    for col in ordered:
        if col in df.columns and col not in seen:
            final_order.append(col)
            seen.add(col)
    for col in df.columns:
        if col not in seen:
            final_order.append(col)
            seen.add(col)

    return df[final_order]


def _format_amount(value):
    return "" if pd.isna(value) else f"S/. {value:,.2f}"


def _format_percent(value):
    return "" if pd.isna(value) else f"{value:.2f}%"


def round_numeric_for_reporting(df):
    """Round monetary/percentage numeric columns to two decimals without altering counts."""
    df = df.copy()
    numeric_cols = df.select_dtypes(include=[np.number]).columns
    for col in numeric_cols:
        if col in EXCLUDE_ROUND_COLS:
            continue
        lower = col.lower()
        if col in PERCENT_DISPLAY_COLUMNS or col.endswith("%"):
            df[col] = df[col].round(2)
        elif any(keyword in lower for keyword in AMOUNT_KEYWORDS):
            df[col] = df[col].round(2)
    return df


def build_style_formatters(df):
    """Return formatter dict for Streamlit Styler with 2-decimal monetary and percent columns."""
    numeric_cols = df.select_dtypes(include=[np.number]).columns
    formatters = {}
    for col in numeric_cols:
        if col in EXCLUDE_ROUND_COLS:
            continue
        lower = col.lower()
        if col in PERCENT_DISPLAY_COLUMNS or col.endswith("%"):
            formatters[col] = _format_percent
        elif any(keyword in lower for keyword in AMOUNT_KEYWORDS):
            formatters[col] = _format_amount
    return formatters


def join_unique_nonempty(values, sep="\n"):
    """Join unique, non-empty string representations preserving order."""

    seen = []
    for value in values:
        if pd.isna(value):
            continue
        text = str(value).strip()
        if not text or text.lower() == "nan":
            continue
        if text not in seen:
            seen.append(text)
    return sep.join(seen)


# =========================
# Utilitarios de carga
# =========================
def autodetect_sheet_and_header(xls, excel_bytes, usecols, user_sheet, header_guess):
    """
    Busca la hoja y la fila que luce como encabezado (contenga 'ano_eje', 'pim', 'pia', 'mto_', 'devenga', 'girado').
    Retorna (sheet_name, header_row_index_pandas).
    """
    candidate_sheets = [user_sheet] if user_sheet else xls.sheet_names
    for s in candidate_sheets:
        try:
            tmp = pd.read_excel(excel_bytes, sheet_name=s, header=None, usecols=usecols, nrows=12)
        except Exception:
            continue
        for r in range(min(8, len(tmp))):
            row_vals = tmp.iloc[r].astype(str).str.lower().tolist()
            hits = sum(int(any(k in v for k in ["ano_eje", "pim", "pia", "mto_", "devenga", "girado"])) for v in row_vals)
            if hits >= 2:
                return s, r
    # Fallback: primera hoja y fila indicada por el usuario - 1 (a √≠ndice 0)
    return xls.sheet_names[0], header_guess - 1


def _flatten_headers(columns):
    """Normaliza encabezados (incluyendo multinivel) en snake_case en min√∫sculas."""

    flattened: List[str] = []
    seen_counts: Dict[str, int] = {}

    for col in columns:
        if isinstance(col, tuple):
            parts: List[str] = []
            for level in col:
                if level is None or (isinstance(level, float) and np.isnan(level)):
                    continue
                text = str(level).strip()
                if not text or text.lower() == "nan":
                    continue
                parts.append(text)
            label = " ".join(parts)
        else:
            if col is None or (isinstance(col, float) and np.isnan(col)):
                label = ""
            else:
                label = str(col).strip()

        if not label:
            label = "col"

        normalized = re.sub(r"\s+", "_", label)
        normalized = normalized.replace("__", "_")
        normalized = normalized.strip("_")
        normalized = normalized.lower() or "col"

        count = seen_counts.get(normalized, 0)
        seen_counts[normalized] = count + 1
        if count:
            normalized = f"{normalized}_{count+1}"

        flattened.append(normalized)

    return flattened


def load_data(excel_bytes, usecols, sheet_name, header_row_excel, autodetect=True):
    xls = pd.ExcelFile(excel_bytes)
    if autodetect:
        s, hdr = autodetect_sheet_and_header(xls, excel_bytes, usecols, sheet_name, header_row_excel)
    else:
        hdr = header_row_excel - 1
        s = sheet_name if sheet_name else xls.sheet_names[0]

    multi_header_df = None
    try:
        multi_header_df = pd.read_excel(
            excel_bytes,
            sheet_name=s,
            header=[hdr, hdr + 1],
            usecols=usecols,
        )
    except Exception:
        pass

    if multi_header_df is not None:
        df = multi_header_df
    else:
        df = pd.read_excel(excel_bytes, sheet_name=s, header=hdr, usecols=usecols)

    df = df.dropna(how="all").dropna(axis=1, how="all")
    df.columns = _flatten_headers(df.columns)
    return df, s

# =========================
# C√°lculos CI‚ÄìEC
# =========================
def find_monthly_columns(df, prefix):
    return [f"{prefix}{i:02d}" for i in range(1, 13) if f"{prefix}{i:02d}" in df.columns]


MONTH_NAME_ALIASES = {
    1: ("1", "01", "ene", "enero", "jan", "january"),
    2: ("2", "02", "feb", "febrero", "febr", "february"),
    3: ("3", "03", "mar", "marzo", "march"),
    4: ("4", "04", "abr", "abril", "april"),
    5: ("5", "05", "may", "mayo"),
    6: ("6", "06", "jun", "junio", "june"),
    7: ("7", "07", "jul", "julio", "july"),
    8: ("8", "08", "ago", "agosto", "aug", "august"),
    9: ("9", "09", "set", "sept", "septiembre", "sep", "september"),
    10: ("10", "oct", "octubre", "october"),
    11: ("11", "nov", "noviembre", "november"),
    12: ("12", "dic", "diciembre", "dec", "december"),
}
MONTH_PROGRAM_COLUMN_PREFERENCES = {
    1: ("ejecutado_01", "pf_enero"),
    2: ("ejecutado_02", "pf_febrero"),
    3: ("ejecutado_03", "pf_marzo"),
    4: ("ejecutado_04", "pf_abril"),
    5: ("ejecutado_05", "pf_mayo"),
    6: ("ejecutado_06", "pf_junio"),
    7: ("ejecutado_07", "pf_julio"),
    8: ("ejecutado_08", "pf_agosto"),
    9: ("ejecutado_09", "pf_setiembre"),
    10: ("pf_octubre", "ejecutado_10"),
    11: ("pf_noviembre", "ejecutado_11"),
    12: ("pf_diciembre", "ejecutado_12"),
}
MONTH_NAME_LABELS = {
    1: "Enero",
    2: "Febrero",
    3: "Marzo",
    4: "Abril",
    5: "Mayo",
    6: "Junio",
    7: "Julio",
    8: "Agosto",
    9: "Setiembre",
    10: "Octubre",
    11: "Noviembre",
    12: "Diciembre",
}
PROGRAM_MATCH_TOKENS = ("prog", "program", "calendario", "cronograma")
PROGRAM_EXCLUDE_TOKENS = ("mto", "pia", "pim", "certificado", "compro", "girado", "devenga")


def _normalize_label(label) -> str:
    if label is None or (isinstance(label, float) and np.isnan(label)):
        return ""
    return re.sub(r"[^a-z0-9]", "", str(label).lower())


def detect_programado_columns(df: pd.DataFrame) -> Dict[int, str]:
    """Infer the monthly programming columns (1-12) from the dataframe headers."""

    normalized_columns: Dict[str, List[str]] = {}
    for col in df.columns:
        normalized_columns.setdefault(_normalize_label(col), []).append(col)

    month_candidates: Dict[int, List[Tuple[str, bool]]] = {i: [] for i in range(1, 13)}
    fallback: List[str] = []

    mapping: Dict[int, str] = {}

    # First, try explicit preferences provided by the SIAF format.
    for month_id, preferences in MONTH_PROGRAM_COLUMN_PREFERENCES.items():
        for candidate in preferences:
            normalized_candidate = _normalize_label(candidate)
            for col in normalized_columns.get(normalized_candidate, []):
                series = df[col]
                if pd.api.types.is_numeric_dtype(series):
                    mapping[month_id] = col
                    break
            if month_id in mapping:
                break

    for col in df.columns:
        series = df[col]
        if not pd.api.types.is_numeric_dtype(series):
            continue
        normalized = _normalize_label(col)
        if not normalized:
            continue
        # Skip columns already mapped via preferences.
        if col in mapping.values():
            continue
        contains_program = any(token in normalized for token in PROGRAM_MATCH_TOKENS)
        month_id = None
        for idx, aliases in MONTH_NAME_ALIASES.items():
            if any(alias in normalized for alias in aliases):
                month_id = idx
                break
        if month_id is not None:
            month_candidates[month_id].append((col, contains_program))
        elif contains_program:
            fallback.append(col)

    for month_id, options in month_candidates.items():
        if month_id in mapping:
            continue
        if not options:
            continue
        options_sorted = sorted(options, key=lambda item: (not item[1], len(str(item[0]))))
        mapping[month_id] = options_sorted[0][0]

    if len(mapping) < 12:
        numeric_columns = [col for col in df.columns if pd.api.types.is_numeric_dtype(df[col])]
        ordered_candidates: List[str] = []
        for col in numeric_columns:
            if col in mapping.values():
                continue
            normalized = _normalize_label(col)
            if any(token in normalized for token in PROGRAM_EXCLUDE_TOKENS):
                continue
            ordered_candidates.append(col)
        for col in fallback:
            if col not in ordered_candidates and col not in mapping.values():
                ordered_candidates.append(col)

        for month_id in range(1, 13):
            if month_id in mapping:
                continue
            if not ordered_candidates:
                break
            mapping[month_id] = ordered_candidates.pop(0)

    return mapping


def attach_programado_metrics(df: pd.DataFrame, month: int):
    """Attach programado_mes and avance_programado_% columns based on detected schedule."""

    df = df.copy()
    month_map = detect_programado_columns(df)
    month_key = int(month) if pd.notna(month) else None
    source_col = month_map.get(month_key) if month_key in month_map else None

    if source_col and source_col in df.columns:
        program_series = pd.to_numeric(df[source_col], errors="coerce").fillna(0.0)
    else:
        program_series = pd.Series(0.0, index=df.index, dtype=float)
        source_col = None

    df["programado_mes"] = program_series.astype(float)
    devengado_mes = pd.to_numeric(df.get("devengado_mes", 0.0), errors="coerce").fillna(0.0).astype(float)
    df["devengado_mes"] = devengado_mes

    program_array = df["programado_mes"].to_numpy(dtype=float, copy=True)
    dev_array = devengado_mes.to_numpy(dtype=float, copy=True)
    ratio = np.zeros_like(program_array)
    np.divide(dev_array, program_array, out=ratio, where=program_array > 0)
    df["avance_programado_%"] = ratio * 100.0

    return df, month_map, source_col


def ensure_ci_ec_steps(df, month, umbral):
    """
    Crea/asegura columnas claves si no existen:
    - devengado (suma mto_devenga_01..12)
    - devengado_mes (columna del mes seleccionado)
    - no_certificado (pim - certificado)
    - avance_% (devengado/pim)
    - riesgo_devolucion (avance_% < umbral)
    - area (vac√≠a si no existe)
    """
    df = df.copy()
    dev_cols = find_monthly_columns(df, "mto_devenga_")

    if "devengado" not in df.columns:
        df["devengado"] = df[dev_cols].sum(axis=1) if dev_cols else 0.0

    col_mes = f"mto_devenga_{int(month):02d}"
    if "devengado_mes" not in df.columns:
        df["devengado_mes"] = df[col_mes] if col_mes in df.columns else 0.0

    pim_series = pd.to_numeric(df.get("mto_pim", 0.0), errors="coerce").fillna(0.0)
    certificado_series = pd.to_numeric(df.get("mto_certificado", 0.0), errors="coerce").fillna(0.0)
    df["no_certificado"] = pim_series - certificado_series

    if "saldo_pim" in df.columns:
        df = df.drop(columns=["saldo_pim"])

    if "avance_%" not in df.columns:
        df["avance_%"] = np.where(df.get("mto_pim", 0) > 0, df["devengado"] / df["mto_pim"] * 100.0, 0.0)

    if "riesgo_devolucion" not in df.columns:
        df["riesgo_devolucion"] = df["avance_%"] < float(umbral)

    if "area" not in df.columns:
        df["area"] = ""

    return df

# =========================
# Clasificador concatenado
# =========================
_code_re = re.compile(r"^\s*(\d+(?:\.\d+)*)")


def extract_code(text):
    """Extrae el prefijo num√©rico (con puntos) de un texto tipo '2.1.1 Bienes y servicios'."""
    if pd.isna(text):
        return ""
    s = str(text).strip()
    m = _code_re.match(s)
    return m.group(1) if m else ""


def last_segment(code):
    return code.split(".")[-1] if code else ""


def concat_hierarchy(gen, sub, subdet, esp, espdet):
    """
    Concatena jer√°rquicamente evitando duplicados:
    generica.subgenerica.subgenerica_det.especifica.especifica_det
    """
    parts = []
    if gen:
        parts.append(gen)
    for child in [sub, subdet, esp, espdet]:
        if not child:
            continue
        if parts and (child.startswith(parts[-1] + ".") or child.startswith(parts[0] + ".")):
            parts.append(child)
        else:
            if parts:
                parts.append(parts[-1] + "." + last_segment(child))
            else:
                parts.append(child)
    return parts[-1] if parts else ""


def normalize_clasificador(code):
    """
    Regla: todo clasificador debe comenzar con '2.'.
    - Si est√° vac√≠o => '2.'
    - Si no inicia con '2.' => anteponer '2.'
    """
    if not code:
        return "2."
    return code if code.startswith("2.") else "2." + code


def desc_only(text):
    """Devuelve solo la descripci√≥n (lo que va despu√©s del primer punto)."""
    if pd.isna(text):
        return ""
    s = str(text)
    return s.split(".", 1)[1].strip() if "." in s else s


def build_classifier_columns(df):
    """
    Crea columnas:
    - gen_cod, sub_cod, subdet_cod, esp_cod, espdet_cod (c√≥digos num√©ricos)
    - clasificador_cod (concatenado y normalizado con 2.)
    - generica_desc, subgenerica_desc, subgenerica_det_desc, especifica_desc, especifica_det_desc
    - clasificador_desc (descripci√≥n jer√°rquica)
    """
    df = df.copy()
    gen = df.get("generica", "")
    sub = df.get("subgenerica", "")
    subdet = df.get("subgenerica_det", "")
    esp = df.get("especifica", "")
    espdet = df.get("especifica_det", "")

    df["gen_cod"] = gen.map(extract_code) if "generica" in df.columns else ""
    df["sub_cod"] = sub.map(extract_code) if "subgenerica" in df.columns else ""
    df["subdet_cod"] = subdet.map(extract_code) if "subgenerica_det" in df.columns else ""
    df["esp_cod"] = esp.map(extract_code) if "especifica" in df.columns else ""
    df["espdet_cod"] = espdet.map(extract_code) if "especifica_det" in df.columns else ""

    df["clasificador_cod"] = [
        normalize_clasificador(concat_hierarchy(g, s, sd, e, ed))
        for g, s, sd, e, ed in zip(
            df["gen_cod"], df["sub_cod"], df["subdet_cod"], df["esp_cod"], df["espdet_cod"]
        )
    ]

    df["generica_desc"] = gen.map(desc_only) if "generica" in df.columns else ""
    df["subgenerica_desc"] = sub.map(desc_only) if "subgenerica" in df.columns else ""
    df["subgenerica_det_desc"] = subdet.map(desc_only) if "subgenerica_det" in df.columns else ""
    df["especifica_desc"] = esp.map(desc_only) if "especifica" in df.columns else ""
    df["especifica_det_desc"] = espdet.map(desc_only) if "especifica_det" in df.columns else ""

    df["clasificador_desc"] = (
        df["generica_desc"].fillna("")
        + " > " + df["subgenerica_desc"].fillna("")
        + " > " + df["subgenerica_det_desc"].fillna("")
        + " > " + df["especifica_desc"].fillna("")
        + " > " + df["especifica_det_desc"].fillna("")
    ).str.strip(" >")

    return df

# =========================
# Pivote / resumen por grupo
# =========================
def pivot_exec(df, group_col, dev_cols):
    cols = []
    if "mto_pia" in df.columns:
        cols.append("mto_pia")
    if "mto_pim" in df.columns:
        cols.append("mto_pim")
    if "mto_certificado" in df.columns:
        cols.append("mto_certificado")
    if "mto_compro_anual" in df.columns:
        cols.append("mto_compro_anual")
    if dev_cols:
        cols.append("devengado")
    if "devengado_mes" in df.columns:
        cols.append("devengado_mes")
    if "programado_mes" in df.columns:
        cols.append("programado_mes")

    if "devengado" not in df.columns and dev_cols:
        df = df.copy()
        df["devengado"] = df[dev_cols].sum(axis=1)

    g = df.groupby(group_col, dropna=False)[cols].sum().reset_index()

    if "mto_pim" in g.columns:
        certificado_col = g["mto_certificado"] if "mto_certificado" in g.columns else 0.0
        g["no_certificado"] = g["mto_pim"] - certificado_col
    if "mto_pim" in g.columns and "devengado" in g.columns:
        g["avance_%"] = np.where(g["mto_pim"] > 0, g["devengado"] / g["mto_pim"] * 100.0, 0.0)
    if "mto_pim" in g.columns and "devengado_mes" in g.columns:
        g["avance_mes_%"] = np.where(g["mto_pim"] > 0, g["devengado_mes"] / g["mto_pim"] * 100.0, 0.0)
    if "programado_mes" in g.columns and "devengado_mes" in g.columns:
        g["avance_programado_%"] = np.where(
            g["programado_mes"] > 0,
            g["devengado_mes"] / g["programado_mes"] * 100.0,
            0.0,
        )

    return g


def _attach_pivot_table(
    workbook_buffer: io.BytesIO,
    source_sheet: str,
    target_sheet: str,
    table_name: str,
    row_fields: List[str],
    value_fields: List[Dict[str, object]],
):
    """Add an Excel pivot table using openpyxl after the workbook is written by XlsxWriter."""

    try:
        from openpyxl import load_workbook
        from openpyxl.pivot.cache import CacheDefinition, CacheField, CacheSource, WorksheetSource, SharedItems
        from openpyxl.pivot.table import (
            DataField,
            Location,
            PivotField,
            PivotTableStyle,
            RowColField,
            RowColItem,
            TableDefinition,
        )
        from openpyxl.utils import get_column_letter
    except Exception:
        return workbook_buffer

    workbook_buffer.seek(0)
    try:
        wb = load_workbook(workbook_buffer)
    except Exception:
        workbook_buffer.seek(0)
        return workbook_buffer

    if source_sheet not in wb.sheetnames:
        workbook_buffer.seek(0)
        return workbook_buffer

    ws_source = wb[source_sheet]
    max_row = ws_source.max_row
    max_col = ws_source.max_column

    if max_row <= 1 or max_col == 0:
        workbook_buffer.seek(0)
        return workbook_buffer

    headers = list(next(ws_source.iter_rows(min_row=1, max_row=1, values_only=True)))
    header_index = {name: idx for idx, name in enumerate(headers) if name}

    if not all(field in header_index for field in row_fields):
        workbook_buffer.seek(0)
        return workbook_buffer

    resolved_values = [vf for vf in value_fields if vf.get("field") in header_index]
    if not resolved_values:
        workbook_buffer.seek(0)
        return workbook_buffer

    if target_sheet in wb.sheetnames:
        del wb[target_sheet]
    ws_pivot = wb.create_sheet(target_sheet)

    data_ref = f"'{source_sheet}'!$A$1:${get_column_letter(max_col)}${max_row}"

    cache_fields = []
    for idx, header in enumerate(headers):
        if header is None:
            continue
        column_cells = list(
            ws_source.iter_cols(
                min_col=idx + 1,
                max_col=idx + 1,
                min_row=2,
                max_row=max_row,
                values_only=True,
            )
        )
        column_values = []
        if column_cells:
            column_values = [cell for cell in column_cells[0] if cell is not None]
        contains_number = any(isinstance(v, (int, float)) for v in column_values)
        contains_string = any(isinstance(v, str) for v in column_values)
        shared_items = SharedItems(
            count=len(column_values),
            containsNumber=contains_number or None,
            containsString=contains_string or None,
            containsBlank=True,
        )
        cache_fields.append(CacheField(name=str(header), sharedItems=shared_items))

    cache = CacheDefinition(
        cacheSource=CacheSource(
            type="worksheet",
            worksheetSource=WorksheetSource(ref=data_ref, sheet=source_sheet),
        ),
        recordCount=max_row - 1,
        cacheFields=tuple(cache_fields),
    )

    cache_id = len(wb._pivots) + 1 or 1
    cache.cacheId = cache_id
    cache._id = cache_id

    pivot_fields = []
    value_field_names = {vf["field"] for vf in resolved_values}
    row_indexes = [header_index[field] for field in row_fields]

    for idx, header in enumerate(headers):
        if header is None:
            continue
        pf = PivotField(name=str(header))
        if idx in row_indexes:
            pf.axis = "axisRow"
            pf.defaultSubtotal = True
        elif header in value_field_names:
            pf.dataField = True
        else:
            pf.defaultSubtotal = False
        pivot_fields.append(pf)

    row_fields_def = [RowColField(x=idx) for idx in row_indexes]
    row_items = [RowColItem(t="grand")]

    data_fields = []
    try:
        from openpyxl.styles.numbers import BUILTIN_FORMATS
    except Exception:
        BUILTIN_FORMATS = {}

    for value in resolved_values:
        field_name = value["field"]
        field_idx = header_index[field_name]
        subtotal = value.get("function", "sum")
        fmt_string = value.get("num_format")
        num_fmt_id = None
        if fmt_string:
            for fmt_key, fmt_val in BUILTIN_FORMATS.items():
                try:
                    matches = fmt_val == fmt_string
                except Exception:
                    matches = False
                if matches:
                    try:
                        num_fmt_id = int(fmt_key)
                    except Exception:
                        num_fmt_id = None
                    if num_fmt_id is not None:
                        break
        df = DataField(
            name=value.get("name", field_name),
            fld=field_idx,
            subtotal=subtotal,
            numFmtId=num_fmt_id,
        )
        data_fields.append(df)

    pivot_style = PivotTableStyle(name="PivotStyleMedium9", showRowHeaders=True, showColHeaders=True, showRowStripes=True)

    pivot = TableDefinition(
        name=table_name,
        cacheId=cache_id,
        dataOnRows=False,
        dataCaption="Valores",
        rowGrandTotals=True,
        colGrandTotals=True,
        location=Location(ref="A3", firstHeaderRow=3, firstDataRow=4, firstDataCol=1),
        pivotFields=tuple(pivot_fields),
        rowFields=tuple(row_fields_def),
        rowItems=tuple(row_items),
        dataFields=tuple(data_fields),
        pivotTableStyleInfo=pivot_style,
    )

    pivot.cache = cache
    ws_pivot.add_pivot(pivot)
    wb._pivots.append(pivot)

    updated_buffer = io.BytesIO()
    wb.save(updated_buffer)
    updated_buffer.seek(0)
    return updated_buffer


def to_excel_download(
    resumen,
    avance,
    proyeccion=None,
    ritmo=None,
    leaderboard=None,
    reporte_siaf=None,
    reporte_siaf_pivot_source=None,
):
    def _populate_workbook(writer: pd.ExcelWriter, use_xlsxwriter: bool):
        workbook = writer.book if use_xlsxwriter else None
        header_format = None
        currency_format = None
        percent_format = None
        if use_xlsxwriter and workbook is not None:
            header_format = workbook.add_format({"bold": True, "bg_color": "#c62828", "font_color": "#ffffff"})
            currency_format = workbook.add_format({"num_format": EXCEL_CURRENCY_FORMAT})
            percent_format = workbook.add_format({"num_format": EXCEL_PERCENT_FORMAT})

        def _sanitize_table_name(name: str) -> str:
            clean = re.sub(r"[^0-9A-Za-z_]", "", name)[:20]
            return clean or "Tabla"

        def add_sheet_with_table(df: pd.DataFrame, sheet_name: str, add_chart: bool = True):
            if df is None or df.empty:
                return None

            df.to_excel(writer, sheet_name=sheet_name, index=False)
            worksheet = writer.sheets[sheet_name]

            max_row, max_col = df.shape
            if use_xlsxwriter and workbook is not None:
                table_name = f"Tbl{_sanitize_table_name(sheet_name)}"
                worksheet.add_table(
                    0,
                    0,
                    max_row,
                    max_col - 1,
                    {
                        "name": table_name,
                        "style": "Table Style Medium 9",
                        "columns": [{"header": col} for col in df.columns],
                    },
                )

                worksheet.set_row(0, None, header_format)

                for col_idx, column_name in enumerate(df.columns):
                    if pd.api.types.is_numeric_dtype(df.iloc[:, col_idx]):
                        if isinstance(column_name, str):
                            if column_name in PERCENT_DISPLAY_COLUMNS:
                                fmt = None
                            elif column_name.endswith("%"):
                                fmt = percent_format
                            else:
                                fmt = currency_format
                        else:
                            fmt = currency_format
                        if fmt is not None:
                            worksheet.set_column(col_idx, col_idx, None, fmt)

                if add_chart and max_row > 0 and max_col > 1:
                    chart = workbook.add_chart({"type": "column"})
                    categories = [sheet_name, 1, 0, max_row, 0]
                    for col_idx in range(1, max_col):
                        if pd.api.types.is_numeric_dtype(df.iloc[:, col_idx]):
                            chart.add_series(
                                {
                                    "name": [sheet_name, 0, col_idx],
                                    "categories": categories,
                                    "values": [sheet_name, 1, col_idx, max_row, col_idx],
                                }
                            )
                    chart.set_title({"name": sheet_name})
                    worksheet.insert_chart(1, max_col + 1, chart, {"x_scale": 1.1, "y_scale": 1.1})

            return worksheet

        add_sheet_with_table(resumen, "Resumen")
        add_sheet_with_table(avance, "Avance")

        if proyeccion is not None and not proyeccion.empty:
            add_sheet_with_table(proyeccion, "Proyeccion")
        if ritmo is not None and not ritmo.empty:
            add_sheet_with_table(ritmo, "Ritmo")
        if leaderboard is not None and not leaderboard.empty:
            add_sheet_with_table(leaderboard, "Leaderboard")

        pivot_source_sheet = None
        pivot_table_config = None
        if reporte_siaf is not None and not reporte_siaf.empty:
            add_sheet_with_table(reporte_siaf, "Reporte_SIAF")

        if reporte_siaf_pivot_source is not None and not reporte_siaf_pivot_source.empty:
            pivot_source_sheet = "Reporte_SIAF_Fuente"
            add_sheet_with_table(reporte_siaf_pivot_source, pivot_source_sheet, add_chart=False)
            pivot_table_config = {
                "rows": ["sec_func", "Generica", "clasificador_cod-concepto"],
                "values": [
                    {"field": "PIM", "function": "sum", "num_format": EXCEL_CURRENCY_FORMAT},
                    {"field": "CERTIFICADO", "function": "sum", "num_format": EXCEL_CURRENCY_FORMAT},
                    {"field": "COMPROMETIDO", "function": "sum", "num_format": EXCEL_CURRENCY_FORMAT},
                    {"field": "DEVENGADO", "function": "sum", "num_format": EXCEL_CURRENCY_FORMAT},
                    {"field": "AVANCE", "function": "average", "num_format": EXCEL_PERCENT_FORMAT},
                    {"field": "DEVENGADO MES", "function": "sum", "num_format": EXCEL_CURRENCY_FORMAT},
                    {"field": "PROGRAMADO MES", "function": "sum", "num_format": EXCEL_CURRENCY_FORMAT},
                    {"field": "AVANCE MES", "function": "average", "num_format": EXCEL_PERCENT_FORMAT},
                    {"field": "NO CERTIFICADO", "function": "sum", "num_format": EXCEL_CURRENCY_FORMAT},
                ],
            }

        return pivot_source_sheet, pivot_table_config

    engine_candidates = []
    missing_modules = set()
    if XLSXWRITER_AVAILABLE:
        engine_candidates.append("xlsxwriter")
    else:
        missing_modules.add("xlsxwriter")
    if OPENPYXL_AVAILABLE:
        engine_candidates.append("openpyxl")
    else:
        missing_modules.add("openpyxl")
    if not engine_candidates:
        missing_summary = ", ".join(sorted(missing_modules)) or "xlsxwriter, openpyxl"
        raise ModuleNotFoundError(
            f"No se encontr√≥ un motor de Excel disponible. Instala {missing_summary}.",
            name=missing_summary,
        )

    pivot_source_sheet = None
    pivot_table_config = None
    output = None
    engine_used = None
    last_exc = None

    for engine in engine_candidates:
        output = io.BytesIO()
        try:
            with pd.ExcelWriter(output, engine=engine) as writer:
                pivot_source_sheet, pivot_table_config = _populate_workbook(writer, engine == "xlsxwriter")
            engine_used = engine
            break
        except ModuleNotFoundError as exc:
            missing_modules.add(getattr(exc, "name", engine))
            last_exc = exc
            continue

    if engine_used is None or output is None:
        missing_summary = ", ".join(sorted(missing_modules)) or "xlsxwriter, openpyxl"
        raise ModuleNotFoundError(
            f"No se encontr√≥ un motor de Excel disponible. Instala {missing_summary}.",
            name=missing_summary,
        ) from last_exc

    output.seek(0)
    if pivot_source_sheet and pivot_table_config:
        output = _attach_pivot_table(
            output,
            pivot_source_sheet,
            "Reporte_SIAF_Pivot",
            "PivotReporteSIAF",
            pivot_table_config["rows"],
            pivot_table_config["values"],
        )
    return output, engine_used

# =========================
# Carga del archivo
# =========================
try:
    df, used_sheet = load_data(
        selected_excel_path,
        usecols,
        sheet_name.strip() or None,
        int(header_row_excel),
        autodetect=detect_header,
    )
except Exception as e:
    st.error(f"No se pudo leer el archivo '{selected_excel_path.name}': {e}")
    st.stop()

st.success(
    f"Le√≠da la hoja '{used_sheet}' del archivo '{selected_excel_path.name}' con {df.shape[0]} filas y {df.shape[1]} columnas."
)

if "sec_func" in df.columns:
    df["sec_func"] = df["sec_func"].apply(map_sec_func)

# =========================
# Filtros
# =========================
st.subheader("Filtros")
filter_cols = [c for c in df.columns if any(k in c for k in [
    "unidad_ejecutora","fuente_financ","generica","especifica_det","funcion",
    "programa_pptal","sec_func","departamento_meta","provincia_meta","area"
])]
filter_cols = [c for c in filter_cols if c not in {"subgenerica", "subgenerica_det"}]

cols_f = st.columns(3)
selected_filters = {}
for i, c in enumerate(filter_cols):
    with cols_f[i % 3]:
        vals = sorted([str(x) for x in df[c].dropna().unique().tolist()])
        if len(vals) > 1:
            pick = st.multiselect(c, options=vals, default=[])
            if pick:
                selected_filters[c] = set(pick)

mask = pd.Series(True, index=df.index)
for c, allowed in selected_filters.items():
    mask &= df[c].astype(str).isin(allowed)
df_f = df[mask].copy()

# =========================
# Aplicar CI‚ÄìEC + Clasificador
# =========================
df_proc = ensure_ci_ec_steps(df_f, current_month, riesgo_umbral)
df_proc = build_classifier_columns(df_proc)
df_proc, program_month_map, program_source_col = attach_programado_metrics(df_proc, current_month)

if program_month_map:
    month_label = MONTH_NAME_LABELS.get(int(current_month), f"Mes {int(current_month):02d}")
    if program_source_col:
        st.caption(
            f"Programaci√≥n del mes {int(current_month):02d} ({month_label}) tomada de la columna "
            f"'{program_source_col}'."
        )
    else:
        st.caption(
            f"No se encontr√≥ columna de programaci√≥n para el mes {int(current_month):02d} ({month_label}); se "
            "asumir√° 0."
        )

    detected_pairs = [
        (MONTH_NAME_LABELS.get(month, f"Mes {month:02d}"), column)
        for month, column in sorted(program_month_map.items())
        if column
    ]
    if detected_pairs:
        items = "".join(
            f"<li><strong>{label}</strong>: <code>{column}</code></li>" for label, column in detected_pairs
        )
        st.markdown(
            f"<small>Columnas detectadas de programaci√≥n mensual:<ul>{items}</ul></small>",
            unsafe_allow_html=True,
        )
else:
    st.caption("No se detectaron columnas de programaci√≥n mensual en el archivo cargado.")

# Totales globales para el resumen ejecutivo
_tot_pia = float(df_proc.get("mto_pia", 0).sum())
_tot_pim = float(df_proc.get("mto_pim", 0).sum())
_tot_dev = float(df_proc.get("devengado", 0).sum())
_tot_cert = float(df_proc.get("mto_certificado", 0).sum()) if "mto_certificado" in df_proc.columns else 0.0
_tot_comp = float(df_proc.get("mto_compro_anual", 0).sum()) if "mto_compro_anual" in df_proc.columns else 0.0
_no_certificado = _tot_pim - _tot_cert
_avance_global = (_tot_dev / _tot_pim * 100.0) if _tot_pim else 0.0

dev_cols = [c for c in df_proc.columns if c.startswith("mto_devenga_")]

_group_options = [c for c in df_proc.columns if c in [
    "clasificador_cod","unidad_ejecutora","fuente_financ","generica","especifica_det",
    "funcion","programa_pptal","sec_func","area"
]]
_group_options = [c for c in _group_options if c not in {"subgenerica", "subgenerica_det"}]
_group_default = _group_options.index("clasificador_cod") if "clasificador_cod" in _group_options else 0
_group_col = st.selectbox("Agrupar por", options=_group_options, index=_group_default)
_group_values = ["(Todos)"] + sorted(df_proc[_group_col].dropna().astype(str).unique().tolist())
_group_filter = st.selectbox(f"Filtrar {_group_col}", options=_group_values, index=0)

df_view = df_proc if _group_filter == "(Todos)" else df_proc[df_proc[_group_col].astype(str) == _group_filter]

# Datos precomputados para cada apartado
pivot = pivot_exec(df_view, _group_col, dev_cols)

_ci_cols = [
    "clasificador_cod", "clasificador_desc",
    "generica","subgenerica","subgenerica_det","especifica","especifica_det",
    "mto_pia","mto_pim","mto_certificado","mto_compro_anual",
    "devengado_mes","programado_mes","devengado","no_certificado",
    "avance_%","avance_programado_%","riesgo_devolucion"
]
_ci_cols = [c for c in _ci_cols if c in df_view.columns]
df_ci = df_view[_ci_cols].head(300) if _ci_cols else pd.DataFrame()

_consol_cols = [
    c
    for c in [
        "mto_pia",
        "mto_pim",
        "mto_certificado",
        "mto_compro_anual",
        "devengado_mes",
        "programado_mes",
        "devengado",
        "no_certificado",
    ]
    if c in df_view.columns
]
consolidado = pd.DataFrame()
if _consol_cols:
    consolidado = df_view.groupby(
        ["clasificador_cod","clasificador_desc","generica","subgenerica","subgenerica_det","especifica","especifica_det"],
        dropna=False
    )[_consol_cols].sum().reset_index()
    if "mto_pim" in consolidado.columns and "devengado" in consolidado.columns:
        consolidado["avance_%"] = np.where(
            consolidado["mto_pim"] > 0,
            consolidado["devengado"] / consolidado["mto_pim"] * 100.0,
            0.0,
        )
    if "programado_mes" in consolidado.columns and "devengado_mes" in consolidado.columns:
        consolidado["avance_programado_%"] = np.where(
            consolidado["programado_mes"] > 0,
            consolidado["devengado_mes"] / consolidado["programado_mes"] * 100.0,
            0.0,
        )

avance_series = pd.DataFrame()
if dev_cols and "mto_pim" in df_view.columns:
    month_map = {f"mto_devenga_{i:02d}": i for i in range(1, 13)}
    dev_series = df_view[dev_cols].sum().reset_index()
    dev_series.columns = ["col", "devengado"]
    dev_series["mes"] = dev_series["col"].map(month_map)
    dev_series = dev_series.dropna(subset=["mes"]).sort_values("mes")
    pim_total = df_view["mto_pim"].sum()
    dev_series["devengado"] = dev_series["devengado"].fillna(0.0)
    dev_series["acumulado"] = dev_series["devengado"].cumsum()
    dev_series["%_acumulado"] = np.where(
        pim_total > 0,
        dev_series["acumulado"] / pim_total * 100.0,
        0.0,
    )
    dev_series["%_acumulado"] = dev_series["%_acumulado"].round(2)
    avance_series = dev_series[["mes", "devengado", "%_acumulado"]]

saldos_monthly = pd.DataFrame()
saldos_monthly_long = pd.DataFrame()
saldos_cumulative_long = pd.DataFrame()
simulation_detail_df = pd.DataFrame()
simulation_overview_df = pd.DataFrame()
simulation_per_gen_df = pd.DataFrame()
simulation_metrics = {
    "baseline_pct": 0.0,
    "intelligent_pct": 0.0,
    "intelligent_return": 0.0,
    "total_pim": 0.0,
    "projected_total": 0.0,
}
if (
    "generica" in df_view.columns
    and program_month_map
    and not df_view.empty
):
    working = df_view.copy()
    working["_generica_label"] = (
        working["generica"].fillna("Sin gen√©rica").astype(str).str.strip()
    )
    working.loc[
        working["_generica_label"].isin(["", "nan", "None"]), "_generica_label"
    ] = "Sin gen√©rica"

    no_certificado_series = pd.to_numeric(
        working.get("no_certificado", 0.0), errors="coerce"
    ).fillna(0.0)
    working["_no_certificado"] = no_certificado_series.astype(float)
    no_cert_by_generica = (
        working.groupby("_generica_label")['_no_certificado'].sum()
        if not working.empty
        else pd.Series(dtype=float)
    )

    month_records: List[Dict[str, float]] = []

    for month in range(1, 13):
        program_col = program_month_map.get(month)
        if not program_col or program_col not in working.columns:
            continue

        dev_col = f"mto_devenga_{month:02d}"
        program_series = pd.to_numeric(
            working[program_col], errors="coerce"
        ).fillna(0.0)
        if dev_col in working.columns:
            dev_series = pd.to_numeric(working[dev_col], errors="coerce").fillna(0.0)
        else:
            dev_series = pd.Series(0.0, index=working.index, dtype=float)

        monthly_frame = pd.DataFrame(
            {
                "_generica_label": working["_generica_label"],
                "_programado": program_series.astype(float),
                "_devengado": dev_series.astype(float),
            }
        )
        aggregated = (
            monthly_frame.groupby("_generica_label")[["_programado", "_devengado"]].sum()
            if not monthly_frame.empty
            else pd.DataFrame(columns=["_programado", "_devengado"])
        )

        for gen_label, row in aggregated.iterrows():
            label = str(gen_label).strip()
            if not label or label.lower() in {"nan", "none"}:
                label = "Sin gen√©rica"

            program_total = float(row["_programado"])
            dev_total = float(row["_devengado"])
            saldo_programado = program_total - dev_total
            month_records.append(
                {
                    "generica": label,
                    "mes": month,
                    "Mes": MONTH_NAME_LABELS.get(month, f"Mes {month:02d}"),
                    "programado": program_total,
                    "devengado": dev_total,
                    "saldo_programado": saldo_programado,
                    "no_certificado": float(no_cert_by_generica.get(gen_label, 0.0)),
                }
            )

    if month_records:
        saldos_monthly = pd.DataFrame.from_records(month_records)
        saldos_monthly = saldos_monthly.sort_values(["generica", "mes"])
        saldos_monthly["saldo_acumulado"] = saldos_monthly.groupby("generica")[
            "saldo_programado"
        ].cumsum()

        saldos_monthly_long = saldos_monthly.melt(
            id_vars=[
                "generica",
                "mes",
                "Mes",
                "programado",
                "devengado",
                "saldo_acumulado",
            ],
            value_vars=["saldo_programado", "no_certificado"],
            var_name="concepto",
            value_name="monto",
        )
        saldos_monthly_long["concepto"] = saldos_monthly_long["concepto"].map(
            {
                "saldo_programado": "Saldo programado (Programado - Devengado)",
                "no_certificado": "No certificado",
            }
        )

        saldos_cumulative_long = saldos_monthly.melt(
            id_vars=["generica", "mes", "Mes"],
            value_vars=["saldo_acumulado", "no_certificado"],
            var_name="concepto",
            value_name="monto",
        )
        saldos_cumulative_long["concepto"] = saldos_cumulative_long["concepto"].map(
            {
                "saldo_acumulado": "Saldo programado acumulado",
                "no_certificado": "No certificado",
            }
        )

if (
    "generica" in df_view.columns
    and "mto_pim" in df_view.columns
    and "devengado" in df_view.columns
    and not df_view.empty
):
    gen_base = df_view.copy()
    gen_base["_generica_label"] = (
        gen_base["generica"].fillna("Sin gen√©rica").astype(str).str.strip()
    )
    gen_base.loc[
        gen_base["_generica_label"].isin(["", "nan", "None"]), "_generica_label"
    ] = "Sin gen√©rica"

    numeric_candidates = [
        "mto_pim",
        "devengado",
        "mto_certificado",
        "no_certificado",
        "programado_mes",
        "devengado_mes",
    ] + dev_cols
    available_numeric = []
    for col in numeric_candidates:
        if col not in gen_base.columns:
            continue
        gen_base[col] = pd.to_numeric(gen_base[col], errors="coerce").fillna(0.0)
        available_numeric.append(col)

    if not available_numeric:
        available_numeric = ["mto_pim", "devengado"]

    aggregated_generica = (
        gen_base.groupby("_generica_label", dropna=False)[available_numeric].sum().reset_index()
    )
    aggregated_generica = aggregated_generica.rename(columns={"_generica_label": "generica"})
    aggregated_generica["generica"] = aggregated_generica["generica"].replace(
        {"": "Sin gen√©rica", "nan": "Sin gen√©rica", "None": "Sin gen√©rica"}
    )

    months_elapsed = max(int(current_month), 1)
    remaining_months = max(12 - months_elapsed, 0)

    aggregated_generica["promedio_mensual"] = (
        aggregated_generica["devengado"] / months_elapsed
    )
    aggregated_generica["proyeccion_restante"] = (
        aggregated_generica["promedio_mensual"] * remaining_months
    )
    aggregated_generica["proyeccion_total"] = (
        aggregated_generica["devengado"] + aggregated_generica["proyeccion_restante"]
    )
    aggregated_generica["proyeccion_total"] = np.minimum(
        aggregated_generica["proyeccion_total"], aggregated_generica["mto_pim"]
    )
    aggregated_generica["saldo_por_devolver"] = (
        aggregated_generica["mto_pim"] - aggregated_generica["proyeccion_total"]
    ).clip(lower=0.0)

    if "no_certificado" not in aggregated_generica.columns:
        aggregated_generica["no_certificado"] = 0.0
    else:
        aggregated_generica["no_certificado"] = aggregated_generica["no_certificado"].clip(lower=0.0)

    eligible_return = np.minimum(
        aggregated_generica["saldo_por_devolver"], aggregated_generica["no_certificado"]
    )
    aggregated_generica["retorno_sugerido"] = np.where(
        eligible_return > 0,
        eligible_return,
        aggregated_generica["saldo_por_devolver"],
    )
    aggregated_generica["retorno_sugerido"] = aggregated_generica["retorno_sugerido"].clip(lower=0.0)

    aggregated_generica["avance_actual_%"] = np.where(
        aggregated_generica["mto_pim"] > 0,
        aggregated_generica["devengado"] / aggregated_generica["mto_pim"] * 100.0,
        0.0,
    )
    aggregated_generica["avance_proyectado_%"] = np.where(
        aggregated_generica["mto_pim"] > 0,
        aggregated_generica["proyeccion_total"] / aggregated_generica["mto_pim"] * 100.0,
        0.0,
    )

    simulation_detail_df = aggregated_generica[
        [
            "generica",
            "mto_pim",
            "devengado",
            "promedio_mensual",
            "proyeccion_total",
            "avance_proyectado_%",
            "saldo_por_devolver",
            "no_certificado",
            "retorno_sugerido",
        ]
    ].copy()

    total_pim = float(simulation_detail_df["mto_pim"].sum())
    projected_total = float(simulation_detail_df["proyeccion_total"].sum())
    baseline_pct = (projected_total / total_pim * 100.0) if total_pim else 0.0

    intelligent_return = float(simulation_detail_df["retorno_sugerido"].sum())
    intelligent_pim = total_pim - intelligent_return
    if intelligent_pim > 0:
        intelligent_projected = min(projected_total, intelligent_pim)
        intelligent_pct = intelligent_projected / intelligent_pim * 100.0
    else:
        intelligent_projected = 0.0
        intelligent_pct = 0.0

    overview_rows = [
        {
            "Escenario": "Sin devoluciones",
            "PIM final": total_pim,
            "Devengado proyectado": projected_total,
            "% avance fin de a√±o": baseline_pct,
        },
        {
            "Escenario": "Devoluci√≥n inteligente",
            "PIM final": max(intelligent_pim, 0.0),
            "Devengado proyectado": intelligent_projected,
            "% avance fin de a√±o": intelligent_pct,
        },
    ]
    simulation_overview_df = pd.DataFrame(overview_rows)

    scenario_rows: List[Dict[str, float]] = []
    for row in aggregated_generica.itertuples():
        if row.retorno_sugerido <= 0:
            continue
        scenario_pim = total_pim - row.retorno_sugerido
        if scenario_pim <= 0:
            continue
        scenario_projected = min(projected_total, scenario_pim)
        scenario_pct = scenario_projected / scenario_pim * 100.0
        scenario_rows.append(
            {
                "generica": row.generica,
                "devolucion": row.retorno_sugerido,
                "pim_final": scenario_pim,
                "%_fin_ano": scenario_pct,
                "delta_pct": scenario_pct - baseline_pct,
            }
        )

    if scenario_rows:
        simulation_per_gen_df = pd.DataFrame(scenario_rows)

    simulation_metrics = {
        "baseline_pct": baseline_pct,
        "baseline_projected": projected_total,
        "intelligent_pct": intelligent_pct,
        "intelligent_return": intelligent_return,
        "intelligent_pim": intelligent_pim,
        "intelligent_projected": intelligent_projected,
        "total_pim": total_pim,
        "projected_total": projected_total,
        "months_elapsed": months_elapsed,
    }

ritmo_df = pd.DataFrame()
leaderboard_df = pd.DataFrame()
reporte_siaf_df = pd.DataFrame()
reporte_siaf_pivot_source = pd.DataFrame()
proyeccion_wide = pd.DataFrame()

# Navegaci√≥n por apartados
(
    tab_resumen,
    tab_consol,
    tab_avance,
    tab_saldos,
    tab_simulacion,
    tab_gestion,
    tab_reporte,
    tab_descarga,
) = st.tabs([
    "Resumen ejecutivo",
    "Consolidado",
    "Avance mensual",
    "Saldos",
    "Simulaciones",
    "Ritmo y alertas",
    "Reporte SIAF",
    "Descargas",
])

with tab_resumen:
    with neon_panel(
        "Resumen ejecutivo",
        icon="üõ°Ô∏è",
        subtitle="Tablero maestro con la telemetr√≠a global del presupuesto",
    ):
        k1, k2, k3, k4, k5, k6, k7 = st.columns(7)
        k1.metric("PIA", f"S/ {_tot_pia:,.2f}")
        k2.metric("PIM", f"S/ {_tot_pim:,.2f}")
        k3.metric("Certificado", f"S/ {_tot_cert:,.2f}")
        k4.metric("Comprometido", f"S/ {_tot_comp:,.2f}")
        k5.metric("Devengado (YTD)", f"S/ {_tot_dev:,.2f}")
        k6.metric("No certificado", f"S/ {_no_certificado:,.2f}")
        k7.metric("Avance", f"{_avance_global:.2f}%")
        pendiente_por_devengar = max(_tot_pim - _tot_dev, 0.0)
        cert_ratio = (_tot_cert / _tot_pim * 100.0) if _tot_pim else 0.0
        render_analysis(
            "El avance global llega a "
            f"{_avance_global:.2f}% del PIM, con S/ {pendiente_por_devengar:,.2f} a√∫n por devengar. "
            f"El certificado cubre el {cert_ratio:.2f}% del presupuesto, dejando S/ {_no_certificado:,.2f} sin certificar.",
            icon="üß≠",
        )

with tab_consol:
    with neon_panel(
        "Consolidado por clasificador",
        icon="üìä",
        subtitle="Top clasificadores con visi√≥n arcade del devengado",
    ):
        if consolidado.empty:
            st.info("No hay informaci√≥n consolidada para mostrar.")
        else:
            consol_display = consolidado.head(500).copy()
            consol_display = standardize_financial_columns(consol_display)
            consol_display = round_numeric_for_reporting(consol_display)
            fmt_consol = build_style_formatters(consol_display)
            consol_style = consol_display.style
            if "AVANCE" in consol_display.columns:
                consol_style = consol_style.applymap(
                    lambda v: "background-color: rgba(255, 41, 109, 0.35); color: #fff;"
                    if v < float(riesgo_umbral)
                    else "",
                    subset=["AVANCE"],
                )
            if fmt_consol:
                consol_style = consol_style.format(fmt_consol)
            st.dataframe(consol_style, use_container_width=True)
            total_clasificadores = int(consolidado.shape[0])
            if "devengado" in consolidado.columns and not consolidado["devengado"].empty:
                dev_series = pd.to_numeric(consolidado["devengado"], errors="coerce").fillna(0.0)
                top_idx = dev_series.idxmax()
                top_row = consolidado.loc[top_idx]
                label_fields = [
                    "clasificador_desc",
                    "clasificador_cod",
                    "generica",
                    "especifica_det",
                ]
                top_label = next(
                    (
                        str(top_row.get(field)).strip()
                        for field in label_fields
                        if field in top_row and str(top_row.get(field)).strip()
                    ),
                    "el clasificador l√≠der",
                )
                top_dev = float(dev_series.loc[top_idx])
                avance_val = top_row.get("avance_%")
                avance_text = (
                    f" con un avance de {float(avance_val):.2f}%"
                    if isinstance(avance_val, (int, float, np.floating)) and not np.isnan(avance_val)
                    else ""
                )
                render_analysis(
                    f"Se resumen {total_clasificadores} clasificadores; {top_label} concentra el mayor devengado "
                    f"con S/ {top_dev:,.2f}{avance_text}.",
                    icon="üí°",
                )
            else:
                render_analysis(
                    f"Se listan {total_clasificadores} clasificadores con montos presupuestales disponibles.",
                    icon="üõ∞Ô∏è",
                )

with tab_avance:
    with neon_panel(
        "Avance mensual interactivo",
        icon="üöÄ",
        subtitle="Comparte ritmo mensual y acumulado como si fuera un marcador de arcade",
    ):
        if avance_series.empty:
            st.info("No hay informaci√≥n de devengado mensual para graficar.")
        else:
            avance_display = avance_series.copy()
            vista_avance = st.radio(
                "Selecciona la vista",
                ("Gr√°fico", "Tabla"),
                horizontal=True,
                key="avance_view_mode",
                help="Alterna entre la visualizaci√≥n gr√°fica y la tabla resumen del devengado mensual.",
                label_visibility="collapsed",
            )

            if vista_avance == "Gr√°fico":
                bar = (
                    alt.Chart(avance_display)
                    .mark_bar(color=ACCENT_COLOR, opacity=0.82)
                    .encode(
                        x=alt.X("mes:O", title="Mes"),
                        y=alt.Y(
                            "devengado:Q",
                            title="Devengado (S/)",
                            axis=alt.Axis(format="$,.2f"),
                        ),
                        tooltip=[
                            alt.Tooltip("mes", title="Mes"),
                            alt.Tooltip("devengado", title="Devengado", format="$,.2f"),
                            alt.Tooltip("%_acumulado", title="% acumulado", format=".2f"),
                        ],
                    )
                )
                line = (
                    alt.Chart(avance_display)
                    .mark_line(color=PRIMARY_COLOR, point=alt.OverlayMarkDef(color="#ffe45c"), strokeWidth=3)
                    .encode(
                        x=alt.X("mes:O", title="Mes"),
                        y=alt.Y(
                            "%_acumulado:Q",
                            title="% acumulado",
                            axis=alt.Axis(format=".2f"),
                        ),
                        tooltip=[
                            alt.Tooltip("mes", title="Mes"),
                            alt.Tooltip("%_acumulado", title="% acumulado", format=".2f"),
                        ],
                    )
                )
                chart = (
                    alt.layer(bar, line)
                    .resolve_scale(y="independent")
                    .properties(width=520, height=280)
                    .interactive()
                )
                st.altair_chart(style_arcade_chart(chart), use_container_width=True)
            else:
                avance_table = round_numeric_for_reporting(avance_display)
                fmt_avance = build_style_formatters(avance_table)
                avance_style = avance_table.style
                if "%_acumulado" in avance_table.columns:
                    avance_style = avance_style.applymap(
                        lambda v: "background-color: rgba(255, 41, 109, 0.35); color: #fff;"
                        if v < float(riesgo_umbral)
                        else "",
                        subset=["%_acumulado"],
                    )
                if fmt_avance:
                    avance_style = avance_style.format(fmt_avance)
                st.dataframe(avance_style, use_container_width=True)
            avance_calc = avance_display.copy()
            if "devengado" in avance_calc.columns:
                avance_calc["devengado"] = pd.to_numeric(avance_calc["devengado"], errors="coerce").fillna(0.0)
                avance_calc = avance_calc.sort_values("mes")
                avance_calc["acumulado"] = avance_calc["devengado"].cumsum()
                last_row = avance_calc.iloc[-1]
                peak_idx = avance_calc["devengado"].idxmax()
                peak_row = avance_calc.loc[peak_idx]
                last_month = int(last_row["mes"]) if not pd.isna(last_row["mes"]) else None
                peak_month = int(peak_row["mes"]) if not pd.isna(peak_row["mes"]) else None
                last_label = (
                    MONTH_NAME_LABELS.get(last_month, f"Mes {last_month:02d}")
                    if last_month
                    else "√öltimo mes"
                )
                peak_label = (
                    MONTH_NAME_LABELS.get(peak_month, f"Mes {peak_month:02d}")
                    if peak_month
                    else "el mes con mayor ejecuci√≥n"
                )
                acumulado = float(last_row.get("acumulado", 0.0))
                avance_pct = float(last_row.get("%_acumulado", 0.0)) if "%_acumulado" in last_row else 0.0
                peak_value = float(peak_row.get("devengado", 0.0))
render_analysis(
    "El devengado acumulado asciende a "
    f"S/ {acumulado:,.2f} a {last_label} ({avance_pct:.2f}% del PIM). "
    f"El mes m√°s din√°mico fue {peak_label} con S/ {peak_value:,.2f} devengados.",
    icon="üìà",
)


with tab_saldos:
    with neon_panel(
        "Saldos programados vs. ejecuci√≥n",
        icon="üõ∞Ô∏è",
        subtitle="Controla los saldos como si fueran barras de energ√≠a por gen√©rica",
    ):
        if saldos_monthly_long.empty or saldos_cumulative_long.empty:
            st.info(
                "No hay informaci√≥n suficiente de programaci√≥n mensual y devengado por gen√©rica para calcular los saldos."
            )
        else:
            genericas_disponibles = sorted(saldos_monthly["generica"].unique().tolist())
            seleccion_genericas = st.multiselect(
                "Gen√©ricas de gasto",
                options=genericas_disponibles,
                default=genericas_disponibles,
                key="saldos_generica_filter",
            )

            if not seleccion_genericas:
                st.warning("Selecciona al menos una gen√©rica para visualizar la evoluci√≥n de saldos.")
            else:
                month_label_order = [
                    MONTH_NAME_LABELS.get(i, f"Mes {i:02d}") for i in range(1, 13)
                ]

                monthly_filtered = saldos_monthly_long[
                    saldos_monthly_long["generica"].isin(seleccion_genericas)
                ]
                cumulative_filtered = saldos_cumulative_long[
                    saldos_cumulative_long["generica"].isin(seleccion_genericas)
                ]

                neon_subheader("Evoluci√≥n mensual de saldos", icon="üìä")
                monthly_chart = (
                    alt.Chart(monthly_filtered)
                    .mark_line(point=alt.OverlayMarkDef(size=70))
                    .encode(
                        x=alt.X("Mes:N", sort=month_label_order, title="Mes"),
                        y=alt.Y("monto:Q", title="Monto (S/)", axis=alt.Axis(format="$,.2f")),
                        color=alt.Color(
                            "generica:N",
                            title="Gen√©rica de gasto",
                            scale=alt.Scale(range=ARCADE_COLOR_RANGE),
                        ),
                        strokeDash=alt.StrokeDash("concepto:N", title="Concepto"),
                        tooltip=[
                            alt.Tooltip("generica", title="Gen√©rica"),
                            alt.Tooltip("Mes", title="Mes"),
                            alt.Tooltip("concepto", title="Concepto"),
                            alt.Tooltip("monto", title="Monto", format="$,.2f"),
                        ],
                    )
                    .properties(height=320)
                )
                st.altair_chart(style_arcade_chart(monthly_chart), use_container_width=True)

                neon_subheader("Saldo acumulado y no certificado", icon="üõ∏")
                cumulative_chart = (
                    alt.Chart(cumulative_filtered)
                    .mark_area(opacity=0.6)
                    .encode(
                        x=alt.X("Mes:N", sort=month_label_order, title="Mes"),
                        y=alt.Y(
                            "monto:Q",
                            stack=False,
                            title="Monto (S/)",
                            axis=alt.Axis(format="$,.2f"),
                        ),
                        color=alt.Color(
                            "concepto:N",
                            title="Concepto",
                            scale=alt.Scale(
                                domain=["Saldo programado acumulado", "No certificado"],
                                range=["#6f2cff", "#ff9d00"],
                            ),
                        ),
                        tooltip=[
                            alt.Tooltip("generica", title="Gen√©rica"),
                            alt.Tooltip("Mes", title="Mes"),
                            alt.Tooltip("concepto", title="Concepto"),
                            alt.Tooltip("monto", title="Monto", format="$,.2f"),
                        ],
                    )
                    .properties(height=320)
                )
                st.altair_chart(style_arcade_chart(cumulative_chart), use_container_width=True)

                month_totals = (
                    monthly_filtered.groupby(["Mes", "concepto"], dropna=False)["monto"].sum().reset_index()
                )
                if not month_totals.empty:
                    hottest = month_totals.sort_values("monto", ascending=False).iloc[0]
                    render_month = hottest["Mes"]
                    render_concept = hottest["concepto"]
                    render_value = float(hottest["monto"])
                else:
                    render_month = "Sin datos"
                    render_concept = "Saldo"
                    render_value = 0.0
                acumulado_total = float(cumulative_filtered["monto"].sum()) if not cumulative_filtered.empty else 0.0
                render_analysis(
                    "En las gen√©ricas seleccionadas, el saldo m√°s desafiante se registr√≥ en "
                    f"{render_month} ({render_concept}) con S/ {render_value:,.2f}. "
                    f"Los montos acumulados suman S/ {acumulado_total:,.2f}, combinando programaci√≥n pendiente y no certificado.",
                    icon="üìü",
                )


with tab_simulacion:
    with neon_panel(
        "Simulaci√≥n de devoluci√≥n de saldos",
        icon="üß†",
        subtitle="Escenarios arcade para optimizar devoluciones y sostener el avance anual",
    ):
        if simulation_detail_df.empty or simulation_overview_df.empty:
            st.info(
                "No hay informaci√≥n suficiente para simular devoluciones; verifica que existan datos de PIM, devengado y saldos por gen√©rica."
            )
        else:
            st.caption(
                "Selecciona autom√°ticamente las combinaciones de devoluci√≥n que reducen el PIM manteniendo la proyecci√≥n de ejecuci√≥n basada en el ritmo mensual."
            )

            detail_display = simulation_detail_df.rename(
                columns={
                    "generica": "Gen√©rica",
                    "mto_pim": "PIM",
                    "devengado": "Devengado acumulado",
                    "promedio_mensual": "Devengado mensual promedio",
                    "proyeccion_total": "Devengado proyectado",
                    "avance_proyectado_%": "% avance proyectado",
                    "saldo_por_devolver": "Saldo por devolver",
                    "no_certificado": "No certificado",
                    "retorno_sugerido": "Devoluci√≥n sugerida",
                }
            )
            detail_display = round_numeric_for_reporting(detail_display)
            fmt_detail = build_style_formatters(detail_display)
            detail_style = detail_display.style
            if fmt_detail:
                detail_style = detail_style.format(fmt_detail)
            neon_subheader("Matriz base de devoluciones", icon="üó∫Ô∏è")
            st.dataframe(detail_style, use_container_width=True)

            adjustable_rows = simulation_detail_df[
                simulation_detail_df["saldo_por_devolver"] > 0
            ].copy()
            custom_returns: Dict[str, float] = {}
            custom_detail_rows: List[Dict[str, float]] = []

            total_pim = simulation_metrics.get("total_pim", 0.0)
            projected_total = simulation_metrics.get("projected_total", 0.0)

            if adjustable_rows.empty:
                st.info(
                    "No hay gen√©ricas con saldo disponible para devolver manualmente."
                )
            else:
                neon_subheader("Panel de control de devoluciones", icon="üéõÔ∏è")
                st.caption(
                    "Arrastra los deslizadores ne√≥n para decidir cu√°nto devolver en cada gen√©rica. El rango m√°ximo se adapta al saldo disponible y parte de la sugerencia inteligente."
                )

                slider_columns = st.columns(min(3, len(adjustable_rows)))
                for idx, row in enumerate(adjustable_rows.itertuples()):
                    gen_value = str(row.generica).strip() or "Gen√©rica"
                    max_available = float(row.saldo_por_devolver)
                    suggested = float(row.retorno_sugerido)
                    if max_available <= 0:
                        continue

                    col = slider_columns[idx % len(slider_columns)]
                    default = min(suggested, max_available)
                    key_suffix = abs(hash(("sim", gen_value))) % 10_000_000
                    slider_key = f"sim_slider_{key_suffix}"
                    saved_value = float(st.session_state.get(slider_key, default))
                    saved_value = float(np.clip(saved_value, 0.0, max_available))
                    step = max(max_available / 40.0, 1.0)

                    with col:
                        st.markdown(f"**{gen_value}**")
                        slider_val = st.slider(
                            "Devoluci√≥n personalizada",
                            min_value=0.0,
                            max_value=float(max_available),
                            value=saved_value,
                            step=float(step),
                            format="S/ %0.0f",
                            key=slider_key,
                        )
                        custom_returns[gen_value] = float(slider_val)
                        st.progress(
                            0.0 if max_available <= 0 else min(slider_val / max_available, 1.0)
                        )
                        st.caption(
                            f"Sugerido: S/ {suggested:,.2f} ¬∑ Disponible: S/ {max_available:,.2f}"
                        )

            overview_rows = simulation_overview_df.to_dict("records")
            custom_return_total = float(sum(custom_returns.values())) if custom_returns else 0.0
            custom_pim_final = max(total_pim - custom_return_total, 0.0)
            if custom_pim_final > 0:
                custom_projected = min(projected_total, custom_pim_final)
                custom_pct = custom_projected / custom_pim_final * 100.0
            else:
                custom_projected = 0.0
                custom_pct = 0.0

            if custom_returns:
                overview_rows.append(
                    {
                        "Escenario": "Devoluci√≥n personalizada",
                        "PIM final": custom_pim_final,
                        "Devengado proyectado": custom_projected,
                        "% avance fin de a√±o": custom_pct,
                    }
                )

            if overview_rows:
                neon_subheader("Marcador hologr√°fico", icon="üí†")
                scoreboard_cols = st.columns(len(overview_rows))
                for idx, record in enumerate(overview_rows):
                    avance = float(record.get("% avance fin de a√±o", 0.0))
                    pim_final = float(record.get("PIM final", 0.0))
                    proyectado = float(record.get("Devengado proyectado", 0.0))
                    delta = proyectado - projected_total if "projected_total" in locals() else proyectado
                    with scoreboard_cols[idx]:
                        st.metric(
                            record.get("Escenario", f"Escenario {idx + 1}"),
                            f"{avance:.2f}%",
                            delta=f"Œî devengado: S/ {delta:,.0f}",
                        )
                        st.caption(
                            f"PIM final: S/ {pim_final:,.0f} ¬∑ Devengado proyectado: S/ {proyectado:,.0f}"
                        )

            for row in simulation_detail_df.itertuples():
                suggested = float(row.retorno_sugerido)
                manual_return = float(custom_returns.get(str(row.generica).strip(), 0.0))
                pim_base = float(row.mto_pim)
                pim_final = max(pim_base - manual_return, 0.0)
                projected = min(float(row.proyeccion_total), pim_final)
                avance_pct = projected / pim_final * 100.0 if pim_final else 0.0
                custom_detail_rows.append(
                    {
                        "generica": row.generica,
                        "pim_base": pim_base,
                        "pim_final": pim_final,
                        "devengado_proyectado": projected,
                        "devolucion_sugerida": suggested,
                        "devolucion_personalizada": manual_return,
                        "%_fin_ano": avance_pct,
                    }
                )

            if custom_returns:
                simulation_metrics.update(
                    {
                        "custom_return": custom_return_total,
                        "custom_pct": custom_pct,
                        "custom_pim": custom_pim_final,
                        "custom_projected": custom_projected,
                    }
                )

            neon_subheader("Tabla t√°ctica de escenarios", icon="üìã")
            overview_df = pd.DataFrame(overview_rows)
            overview_display = round_numeric_for_reporting(overview_df.copy())
            fmt_overview = build_style_formatters(overview_display)
            overview_style = overview_display.style
            if fmt_overview:
                overview_style = overview_style.format(fmt_overview)
            st.dataframe(overview_style, use_container_width=True)

            neon_subheader("Arena interactiva de escenarios", icon="üïπÔ∏è")
            st.caption(
                "Activa el tablero hologr√°fico para comparar c√≥mo evoluciona el devengado hist√≥rico y la proyecci√≥n futura en cada escenario."
            )
            month_label_order = [
                MONTH_NAME_LABELS.get(i, f"Mes {i:02d}") for i in range(1, 13)
            ]
            months_elapsed = max(int(current_month), 1)
            monthly_actual_map: Dict[int, float] = {}
            if not avance_series.empty:
                for row in avance_series.itertuples():
                    month_val = getattr(row, "mes", None)
                    if pd.isna(month_val):
                        continue
                    dev_val = float(getattr(row, "devengado", 0.0) or 0.0)
                    monthly_actual_map[int(month_val)] = dev_val
            executed_total = float(simulation_detail_df["devengado"].sum()) if not simulation_detail_df.empty else 0.0
            if not monthly_actual_map and months_elapsed > 0:
                even_value = executed_total / months_elapsed if months_elapsed else 0.0
                monthly_actual_map = {m: even_value for m in range(1, months_elapsed + 1)}
            for m in range(1, months_elapsed + 1):
                monthly_actual_map.setdefault(m, 0.0)

            scenario_series_map: Dict[str, pd.DataFrame] = {}

            def _build_scenario_series(name: str, pim_final: float, projected_total: float) -> pd.DataFrame:
                pim_final = float(pim_final or 0.0)
                projected_total = float(projected_total or 0.0)
                rows: List[Dict[str, float]] = []
                cumulative = 0.0
                historical_total = 0.0
                for month in range(1, months_elapsed + 1):
                    value = float(monthly_actual_map.get(month, 0.0))
                    historical_total += value
                    cumulative += value
                    rows.append(
                        {
                            "Escenario": name,
                            "Mes": MONTH_NAME_LABELS.get(month, f"Mes {month:02d}"),
                            "MesIndex": month,
                            "Tipo": "Hist√≥rico",
                            "Devengado": value,
                            "Acumulado": cumulative,
                            "Avance_pct": (cumulative / pim_final * 100.0) if pim_final > 0 else 0.0,
                        }
                    )
                future_months = list(range(months_elapsed + 1, 13))
                future_total = max(projected_total - historical_total, 0.0)
                if future_months:
                    if rows:
                        last_hist = rows[-1].copy()
                        last_hist["Tipo"] = "Proyecci√≥n"
                        last_hist["Devengado"] = 0.0
                        rows.append(last_hist)
                    if future_total <= 0:
                        future_values = [0.0] * len(future_months)
                    else:
                        base_value = future_total / len(future_months)
                        future_values = [base_value] * len(future_months)
                        adjustment = future_total - sum(future_values)
                        if future_values:
                            future_values[-1] += adjustment
                        future_values = [max(v, 0.0) for v in future_values]
                    for month, value in zip(future_months, future_values):
                        cumulative += float(value)
                        rows.append(
                            {
                                "Escenario": name,
                                "Mes": MONTH_NAME_LABELS.get(month, f"Mes {month:02d}"),
                                "MesIndex": month,
                                "Tipo": "Proyecci√≥n",
                                "Devengado": float(value),
                                "Acumulado": cumulative,
                                "Avance_pct": (cumulative / pim_final * 100.0) if pim_final > 0 else 0.0,
                            }
                        )
                return pd.DataFrame(rows)

            for record in overview_rows:
                name = str(record.get("Escenario", "Escenario")).strip()
                pim_final = float(record.get("PIM final", 0.0))
                projected_total = float(record.get("Devengado proyectado", 0.0))
                scenario_df = _build_scenario_series(name, pim_final, projected_total)
                if not scenario_df.empty:
                    scenario_series_map[name] = scenario_df

            if scenario_series_map:
                scenario_names = list(scenario_series_map.keys())
                selected_scenario = st.radio(
                    "Selecciona el escenario a proyectar",
                    scenario_names,
                    horizontal=True,
                    key="arena_scenario_selector",
                )
                selected_df = scenario_series_map[selected_scenario].sort_values("MesIndex")
                trajectory_chart = (
                    alt.Chart(selected_df)
                    .mark_line(strokeWidth=4)
                    .encode(
                        x=alt.X("Mes:N", sort=month_label_order, title="Mes"),
                        y=alt.Y(
                            "Acumulado:Q",
                            title="Devengado acumulado (S/)",
                            axis=alt.Axis(format="$,.2f"),
                        ),
                        color=alt.Color(
                            "Tipo:N",
                            title="Tramo",
                            scale=alt.Scale(domain=["Hist√≥rico", "Proyecci√≥n"], range=["#00eaff", "#ff296d"]),
                        ),
                        tooltip=[
                            alt.Tooltip("Mes", title="Mes"),
                            alt.Tooltip("Tipo", title="Tramo"),
                            alt.Tooltip("Acumulado", title="Devengado acumulado", format="$,.2f"),
                            alt.Tooltip("Avance_pct", title="Avance (%)", format=".2f"),
                        ],
                        detail="Escenario:N",
                    )
                    .properties(height=320)
                )
                trajectory_points = (
                    alt.Chart(selected_df)
                    .mark_point(filled=True, size=100, color="#ffe45c")
                    .encode(
                        x=alt.X("Mes:N", sort=month_label_order),
                        y=alt.Y("Acumulado:Q"),
                        tooltip=[
                            alt.Tooltip("Mes", title="Mes"),
                            alt.Tooltip("Tipo", title="Tramo"),
                            alt.Tooltip("Devengado", title="Devengado del mes", format="$,.2f"),
                            alt.Tooltip("Acumulado", title="Devengado acumulado", format="$,.2f"),
                            alt.Tooltip("Avance_pct", title="Avance (%)", format=".2f"),
                        ],
                    )
                )
                st.altair_chart(
                    style_arcade_chart(trajectory_chart + trajectory_points),
                    use_container_width=True,
                )

                avance_chart = (
                    alt.Chart(selected_df)
                    .mark_line(strokeWidth=3)
                    .encode(
                        x=alt.X("Mes:N", sort=month_label_order, title="Mes"),
                        y=alt.Y(
                            "Avance_pct:Q",
                            title="Avance acumulado (%)",
                            axis=alt.Axis(format=".2f"),
                        ),
                        color=alt.Color(
                            "Tipo:N",
                            title="Tramo",
                            scale=alt.Scale(domain=["Hist√≥rico", "Proyecci√≥n"], range=["#19f5aa", "#ff9d00"]),
                        ),
                        tooltip=[
                            alt.Tooltip("Mes", title="Mes"),
                            alt.Tooltip("Tipo", title="Tramo"),
                            alt.Tooltip("Avance_pct", title="Avance (%)", format=".2f"),
                        ],
                        detail="Escenario:N",
                    )
                    .properties(height=220)
                )
                st.altair_chart(style_arcade_chart(avance_chart), use_container_width=True)
            else:
                st.info(
                    "No se pudo generar la l√≠nea hist√≥rica y proyectada porque no hay datos mensuales de devengado disponibles."
                )

            if custom_returns:
                chart_return_rows = []
                for row in adjustable_rows.itertuples():
                    chart_return_rows.append(
                        {
                            "Gen√©rica": row.generica,
                            "Tipo": "Sugerido",
                            "Monto": float(row.retorno_sugerido),
                        }
                    )
                    chart_return_rows.append(
                        {
                            "Gen√©rica": row.generica,
                            "Tipo": "Personalizado",
                            "Monto": float(custom_returns.get(str(row.generica).strip(), 0.0)),
                        }
                    )
                returns_chart_df = pd.DataFrame(chart_return_rows)
                if not returns_chart_df.empty:
                    hover_selection = alt.selection_single(fields=["Gen√©rica"], nearest=True, on="mouseover")
                    returns_chart = (
                        alt.Chart(returns_chart_df)
                        .add_selection(hover_selection)
                        .mark_bar(cornerRadiusTopLeft=12, cornerRadiusTopRight=12)
                        .encode(
                            x=alt.X("Gen√©rica:N", sort="-y", title="Gen√©rica"),
                            y=alt.Y("Monto:Q", title="Monto (S/)", axis=alt.Axis(format="$,.2f")),
                            color=alt.Color(
                                "Tipo:N",
                                title="Escenario",
                                scale=alt.Scale(range=["#00eaff", "#ff296d"]),
                            ),
                            opacity=alt.condition(hover_selection, alt.value(1.0), alt.value(0.45)),
                            tooltip=[
                                alt.Tooltip("Gen√©rica:N", title="Gen√©rica"),
                                alt.Tooltip("Tipo:N", title="Escenario"),
                                alt.Tooltip("Monto:Q", title="Monto", format="$,.2f"),
                            ],
                        )
                        .properties(height=320)
                        .interactive()
                    )
                    st.altair_chart(style_arcade_chart(returns_chart), use_container_width=True)

            if not simulation_per_gen_df.empty:
                neon_subheader("Impacto por gen√©rica evaluada", icon="üõ∞Ô∏è")
                per_gen_display = simulation_per_gen_df.rename(
                    columns={
                        "generica": "Gen√©rica",
                        "devolucion": "Devoluci√≥n evaluada",
                        "pim_final": "PIM final",
                        "%_fin_ano": "% avance fin de a√±o",
                        "delta_pct": "Variaci√≥n vs base (p.p.)",
                    }
                )
                per_gen_display = round_numeric_for_reporting(per_gen_display)
                fmt_per_gen = build_style_formatters(per_gen_display)
                per_gen_style = per_gen_display.style
                if fmt_per_gen:
                    per_gen_style = per_gen_style.format(fmt_per_gen)
                st.dataframe(per_gen_style, use_container_width=True)

            if custom_returns and custom_detail_rows:
                custom_per_gen_df = pd.DataFrame(custom_detail_rows)
                custom_per_gen_df = custom_per_gen_df[
                    custom_per_gen_df["devolucion_personalizada"] > 0
                ]
                if not custom_per_gen_df.empty:
                    neon_subheader("Impacto personalizado por gen√©rica", icon="üåå")
                    custom_display = custom_per_gen_df.rename(
                        columns={
                            "generica": "Gen√©rica",
                            "pim_base": "PIM base",
                            "pim_final": "PIM final",
                            "devengado_proyectado": "Devengado proyectado",
                            "devolucion_sugerida": "Devoluci√≥n sugerida",
                            "devolucion_personalizada": "Devoluci√≥n personalizada",
                            "%_fin_ano": "% avance fin de a√±o",
                        }
                    )
                    custom_display = round_numeric_for_reporting(custom_display)
                    fmt_custom = build_style_formatters(custom_display)
                    custom_style = custom_display.style
                    if fmt_custom:
                        custom_style = custom_style.format(fmt_custom)
                    st.dataframe(custom_style, use_container_width=True)

            baseline_pct = simulation_metrics.get("baseline_pct", 0.0)
            intelligent_pct = simulation_metrics.get("intelligent_pct", baseline_pct)
            intelligent_return = simulation_metrics.get("intelligent_return", 0.0)
            projected_total = simulation_metrics.get("projected_total", 0.0)
            total_pim = simulation_metrics.get("total_pim", 0.0)
            custom_pct = simulation_metrics.get("custom_pct")
            custom_return_total = simulation_metrics.get("custom_return", 0.0)
            custom_projected = simulation_metrics.get("custom_projected")
            month_label = MONTH_NAME_LABELS.get(int(current_month), f"Mes {int(current_month):02d}")

            top_candidates = simulation_detail_df[simulation_detail_df["retorno_sugerido"] > 0]
            top_candidates = top_candidates.sort_values("retorno_sugerido", ascending=False).head(3)
            if intelligent_return > 0 and not top_candidates.empty:
                bullets = "; ".join(
                    f"{row.generica}: S/ {row.retorno_sugerido:,.2f}" for row in top_candidates.itertuples()
                )
                analysis_text = (
                    f"Con el ritmo promedio observado hasta {month_label} se proyecta un avance de {baseline_pct:.2f}% sobre un PIM de S/ {total_pim:,.2f}. "
                    f"Devolver inteligentemente S/ {intelligent_return:,.2f} concentrados en {bullets} elevar√≠a la proyecci√≥n a {intelligent_pct:.2f}% "
                    f"manteniendo un devengado estimado de S/ {projected_total:,.2f}."
                )
            elif intelligent_return > 0:
                analysis_text = (
                    f"El algoritmo recomienda devolver S/ {intelligent_return:,.2f} para elevar el avance esperado de {baseline_pct:.2f}% a {intelligent_pct:.2f}% sin reducir el devengado proyectado."
                )
            else:
                analysis_text = (
                    f"Con el ritmo actual se espera ejecutar S/ {projected_total:,.2f} ({baseline_pct:.2f}% del PIM); no se proyectan saldos sobrantes que ameriten devoluci√≥n."
                )

            if (
                custom_returns
                and custom_pct is not None
                and custom_projected is not None
                and custom_return_total > 0
            ):
                analysis_text += (
                    f" Al aplicar la devoluci√≥n personalizada de S/ {custom_return_total:,.2f} el avance estimado alcanzar√≠a {custom_pct:.2f}% con un devengado proyectado de S/ {custom_projected:,.2f}."
                )

            if not simulation_per_gen_df.empty:
                best_delta = simulation_per_gen_df.sort_values("delta_pct", ascending=False).iloc[0]
                if best_delta["delta_pct"] > 0:
                    analysis_text += (
                        f" Adem√°s, devolver √∫nicamente en {best_delta['generica']} implicar√≠a un avance estimado de {best_delta['%_fin_ano']:.2f}% (+{best_delta['delta_pct']:.2f} p.p. frente al escenario base)."
                    )

            render_analysis(analysis_text, icon="üß†")

with tab_gestion:
    with neon_panel(
        "Ritmo requerido por proceso",
        icon="‚öôÔ∏è",
        subtitle="Usa la consola de potencia para alinear cada proceso con el objetivo anual",
    ):
        if "mto_pim" not in df_view.columns:
            st.info("No hay datos de PIM para calcular el ritmo requerido.")
        else:
            remaining_months = max(12 - current_month, 1)
            pim_total = df_view["mto_pim"].sum()
            processes = []
            for col, label in [
                ("mto_certificado", "Certificar"),
                ("mto_compro_anual", "Comprometer"),
                ("devengado", "Devengar"),
            ]:
                total = df_view.get(col, pd.Series(dtype=float)).sum()
                actual_avg = total / max(current_month, 1)
                needed = max(pim_total - total, 0)
                required_avg = needed / remaining_months
                processes.append({"Proceso": label, "Actual": actual_avg, "Necesario": required_avg})
            ritmo_raw = pd.DataFrame(processes)
            if ritmo_raw.empty:
                st.info("No hay informaci√≥n suficiente para calcular el ritmo requerido.")
            else:
                st.caption(
                    "Activa la consola interactiva para impulsar cada proceso. Ajusta los deslizadores y observa c√≥mo responde el tablero en tiempo real."
                )
                impulse_values: List[float] = []
                control_columns = st.columns(len(ritmo_raw))
                for idx, row in ritmo_raw.iterrows():
                    proceso = str(row["Proceso"])
                    actual = float(row["Actual"])
                    necesario = float(row["Necesario"])
                    gap = max(necesario - actual, 0.0)
                    max_slider = float(max(necesario * 1.5, actual * 1.5, gap * 2.0, 1.0))
                    key_suffix = abs(hash(("ritmo", proceso))) % 10_000_000
                    slider_key = f"ritmo_impulso_{key_suffix}"
                    saved_value = float(st.session_state.get(slider_key, gap))
                    saved_value = float(np.clip(saved_value, 0.0, max_slider))
                    step = max(max_slider / 40.0, 1.0)

                    col = control_columns[idx % len(control_columns)]
                    with col:
                        st.markdown(f"**{proceso}**")
                        st.metric("Ritmo actual", f"S/ {actual:,.2f}", delta=f"Meta: S/ {necesario:,.2f}")
                        progress_ratio = 1.0 if necesario <= 0 else min(actual / necesario, 1.0)
                        st.progress(progress_ratio)
                        slider_val = st.slider(
                            "Impulso mensual",
                            min_value=0.0,
                            max_value=max_slider,
                            value=saved_value,
                            step=float(step),
                            format="S/ %0.0f",
                            key=slider_key,
                        )
                        impulse_values.append(float(slider_val))
                        st.caption(f"Brecha actual: S/ {gap:,.2f}")

                if len(impulse_values) < len(ritmo_raw):
                    impulse_values.extend([0.0] * (len(ritmo_raw) - len(impulse_values)))

                ritmo_dynamic = ritmo_raw.copy()
                ritmo_dynamic["Impulso manual (S/)"] = impulse_values
                ritmo_dynamic["Ritmo ajustado"] = ritmo_dynamic["Actual"] + ritmo_dynamic["Impulso manual (S/)"]
                ritmo_dynamic["Brecha restante (S/)"] = (
                    ritmo_dynamic["Necesario"] - ritmo_dynamic["Ritmo ajustado"]
                )

                ritmo_display = round_numeric_for_reporting(
                    ritmo_dynamic.rename(
                        columns={
                            "Actual": "Ritmo actual (S/)",
                            "Necesario": "Ritmo necesario (S/)",
                        }
                    )
                )
                fmt_ritmo = build_style_formatters(ritmo_display)
                ritmo_style = ritmo_display.style
                if fmt_ritmo:
                    ritmo_style = ritmo_style.format(fmt_ritmo)
                st.dataframe(ritmo_style, use_container_width=True)

                neon_subheader("Tablero visual ne√≥n", icon="üßæ")
                ritmo_melt = ritmo_dynamic.melt(
                    "Proceso",
                    value_vars=["Actual", "Necesario", "Ritmo ajustado"],
                    var_name="Escenario",
                    value_name="Monto",
                )
                ritmo_chart = (
                    alt.Chart(ritmo_melt)
                    .mark_bar(cornerRadiusTopLeft=12, cornerRadiusTopRight=12)
                    .encode(
                        x=alt.X("Proceso:N", title="Proceso"),
                        y=alt.Y("Monto:Q", title="Monto mensual (S/)", axis=alt.Axis(format="$,.2f")),
                        color=alt.Color(
                            "Escenario:N",
                            title="Escenario",
                            scale=alt.Scale(range=["#00eaff", "#ff296d", "#ffe45c"]),
                        ),
                        tooltip=[
                            alt.Tooltip("Proceso:N", title="Proceso"),
                            alt.Tooltip("Escenario:N", title="Escenario"),
                            alt.Tooltip("Monto:Q", title="Monto", format="$,.2f"),
                        ],
                    )
                    .properties(height=320)
                    .interactive()
                )
                st.altair_chart(style_arcade_chart(ritmo_chart), use_container_width=True)

                brecha_chart_df = ritmo_dynamic.copy()
                brecha_chart_df["Brecha positiva"] = brecha_chart_df["Brecha restante (S/)"]
                brecha_chart_df["Brecha positiva"] = brecha_chart_df["Brecha positiva"].clip(lower=0.0)
                if brecha_chart_df["Brecha positiva"].sum() > 0:
                    brecha_chart = (
                        alt.Chart(brecha_chart_df)
                        .mark_area(line={"color": "#ff296d", "size": 3})
                        .encode(
                            x=alt.X("Proceso:N", title="Proceso"),
                            y=alt.Y("Brecha positiva:Q", title="Brecha pendiente (S/)", axis=alt.Axis(format="$,.2f")),
                            color=alt.value("#ff296d"),
                            tooltip=[
                                alt.Tooltip("Proceso:N", title="Proceso"),
                                alt.Tooltip("Brecha positiva:Q", title="Brecha", format="$,.2f"),
                            ],
                        )
                        .properties(height=220)
                        .interactive()
                    )
                    st.altair_chart(style_arcade_chart(brecha_chart), use_container_width=True)

                if not ritmo_dynamic.empty:
                    brechas = ritmo_dynamic["Brecha restante (S/)"]
                    brechas_clipped = brechas.clip(lower=0.0)
                    total_brecha = float(brechas_clipped.sum())
                    if total_brecha <= 1e-6:
                        render_analysis(
                            "Con el impulso definido cada proceso alcanza el ritmo necesario; el programa se mantiene al d√≠a con el PIM proyectado.",
                            icon="‚úÖ",
                        )
                    else:
                        peor_idx = brechas_clipped.idxmax()
                        proceso_objetivo = ritmo_dynamic.loc[peor_idx, "Proceso"]
                        brecha = float(brechas_clipped.loc[peor_idx])
                        requerido_total = float(ritmo_dynamic["Necesario"].sum())
                        render_analysis(
                            "Para completar el PIM se requiere ejecutar en promedio S/ "
                            f"{requerido_total:,.2f} mensuales. Tras el impulso propuesto a√∫n falta potenciar {proceso_objetivo} "
                            f"en S/ {brecha:,.2f} mensuales.",
                            icon="‚ö†Ô∏è",
                        )

    with neon_panel(
        "Top √°reas con menor avance",
        icon="üèÅ",
        subtitle="Ranking din√°mico de las √°reas rezagadas con indicadores cr√≠ticos",
    ):
        if "sec_func" in df_view.columns and "mto_pim" in df_view.columns:
            agg_cols = ["mto_pim", "devengado", "devengado_mes", "programado_mes"]
            if "mto_certificado" in df_view.columns:
                agg_cols.insert(1, "mto_certificado")
            agg_sec = df_view.groupby("sec_func", dropna=False)[agg_cols].sum().reset_index()
            agg_sec = agg_sec[agg_sec["mto_pim"] > 0].copy()
            if agg_sec.empty:
                st.info("No hay √°reas con PIM positivo para calcular el rendimiento.")
            else:
                agg_sec["avance_acum_%"] = np.where(
                    agg_sec["mto_pim"] > 0, agg_sec["devengado"] / agg_sec["mto_pim"] * 100.0, 0.0
                )
                agg_sec["avance_mes_%"] = np.where(
                    agg_sec["mto_pim"] > 0, agg_sec["devengado_mes"] / agg_sec["mto_pim"] * 100.0, 0.0,
                )
                agg_sec["avance_programado_%"] = np.where(
                    agg_sec["programado_mes"] > 0,
                    agg_sec["devengado_mes"] / agg_sec["programado_mes"] * 100.0,
                    0.0,
                )
                agg_sec["rank_acum"] = agg_sec["avance_acum_%"].rank(method="dense", ascending=True).astype(int)
                agg_sec["rank_mes"] = agg_sec["avance_mes_%"].rank(method="dense", ascending=True).astype(int)

                max_top = max(int(agg_sec.shape[0]), 1)
                top_default = min(5, max_top)
                slider_key = "leaderboard_top_n"
                slider_value = st.session_state.get(slider_key, top_default)
                if slider_value < 1 or slider_value > max_top:
                    slider_value = min(max(slider_value, 1), max_top)
                    st.session_state[slider_key] = slider_value

                top_n = st.slider(
                    "N√∫mero de √°reas a mostrar",
                    min_value=1,
                    max_value=max_top,
                    value=slider_value,
                    key=slider_key,
                )

                leaderboard_df = (
                    agg_sec.sort_values(["avance_acum_%", "avance_mes_%"], ascending=[True, True])
                    .head(top_n)
                    .copy()
                )
                display_cols = ["rank_acum", "rank_mes", "sec_func", "mto_pim"]
                if "mto_certificado" in agg_sec.columns:
                    display_cols.append("mto_certificado")
                display_cols.extend([
                    "devengado",
                    "avance_acum_%",
                    "devengado_mes",
                    "programado_mes",
                    "avance_mes_%",
                    "avance_programado_%",
                ])
                leaderboard_df = leaderboard_df[display_cols]

                leaderboard_display = leaderboard_df.copy()
                leaderboard_display = standardize_financial_columns(leaderboard_display)
                leaderboard_display = round_numeric_for_reporting(leaderboard_display)
                fmt_leader = build_style_formatters(leaderboard_display)
                leader_style = leaderboard_display.style.applymap(
                    lambda v: "background-color: rgba(255, 41, 109, 0.35); color: #fff;"
                    if v < float(riesgo_umbral)
                    else "",
                    subset=[
                        c
                        for c in ["AVANCE", "AVANCE MES", "AVANCE MES (PIM)"]
                        if c in leaderboard_display.columns
                    ],
                )
                if fmt_leader:
                    leader_style = leader_style.format(fmt_leader)
                st.dataframe(leader_style, use_container_width=True)
                if not leaderboard_df.empty:
                    worst_row = leaderboard_df.iloc[0]
                    sec_label = str(worst_row.get("sec_func", "El primer sec_func")).strip()
                    avance_acum = float(worst_row.get("avance_acum_%", 0.0))
                    avance_mes = float(worst_row.get("avance_mes_%", 0.0))
                    programado_mes = float(worst_row.get("programado_mes", 0.0))
                    devengado_mes = float(worst_row.get("devengado_mes", 0.0))
                    render_analysis(
                        f"El √°rea {sec_label} lidera la zona cr√≠tica con un avance acumulado de {avance_acum:.2f}% y un avance mensual de {avance_mes:.2f}%. "
                        f"En el mes ejecut√≥ S/ {devengado_mes:,.2f} frente a un programado de S/ {programado_mes:,.2f}.",
                        icon="üö®",
                    )
        else:
            st.info("No hay informaci√≥n de sec_func con PIM positivo para generar el ranking.")

with tab_reporte:
    with neon_panel(
        "Reporte SIAF por √°rea, gen√©rica y espec√≠fica detalle",
        icon="üìö",
        subtitle="Explora la jerarqu√≠a de clasificadores con el brillo de una interfaz arcade",
    ):
        reporte_siaf_pivot_source = pd.DataFrame()
        if not all(col in df_view.columns for col in ["sec_func", "generica", "especifica_det"]):
            st.info("Para el reporte SIAF se requieren las columnas sec_func, generica y especifica_det.")
        else:
            siaf_agg_cols = [
                c
                for c in [
                    "mto_pia",
                    "mto_pim",
                    "mto_certificado",
                    "mto_compro_anual",
                    "devengado_mes",
                    "programado_mes",
                    "devengado",
                    "no_certificado",
                ]
                if c in df_view.columns
            ]

            if not siaf_agg_cols:
                st.info("No se encontraron columnas monetarias para generar el reporte SIAF por √°rea.")
            else:
                base_group = ["sec_func", "generica", "especifica_det"]
                agg_map = {col: "sum" for col in siaf_agg_cols}
                if "clasificador_cod" in df_view.columns:
                    agg_map["clasificador_cod"] = "first"
                if "especifica_det_desc" in df_view.columns:
                    agg_map["especifica_det_desc"] = "first"

                reporte_base = (
                    df_view.groupby(base_group, dropna=False)
                    .agg(agg_map)
                    .reset_index()
                )

            if "clasificador_cod" in reporte_base.columns:
                clasificador_cod = reporte_base["clasificador_cod"].fillna("").astype(str).str.strip()
            else:
                clasificador_cod = reporte_base["especifica_det"].map(extract_code).fillna("").astype(str).str.strip()

            if "especifica_det_desc" in reporte_base.columns:
                concepto = reporte_base["especifica_det_desc"].fillna("").astype(str).str.strip()
            else:
                concepto = reporte_base["especifica_det"].map(desc_only).fillna("").astype(str).str.strip()

            tiene_cod = (clasificador_cod != "") & (clasificador_cod != "nan")
            tiene_concepto = (concepto != "") & (concepto != "nan")
            reporte_base["clasificador_cod_concepto"] = np.where(
                tiene_cod & tiene_concepto,
                clasificador_cod + " - " + concepto,
                np.where(tiene_concepto, concepto, clasificador_cod),
            )

            value_sources = {
                "PIM": "mto_pim",
                "CERTIFICADO": "mto_certificado",
                "COMPROMETIDO": "mto_compro_anual",
                "DEVENGADO": "devengado",
                "DEVENGADO MES": "devengado_mes",
                "PROGRAMADO MES": "programado_mes",
                "NO CERTIFICADO": "no_certificado",
            }
            for src in value_sources.values():
                if src not in reporte_base.columns:
                    reporte_base[src] = 0.0

            reporte_base = reporte_base[
                reporte_base["clasificador_cod_concepto"].fillna("").astype(str).str.strip() != ""
            ].copy()

            if not reporte_base.empty:
                def _safe_numeric(col_name):
                    if col_name in reporte_base.columns:
                        return reporte_base[col_name].fillna(0.0).astype(float)
                    return pd.Series(0.0, index=reporte_base.index, dtype=float)

                devengado_acum = _safe_numeric("devengado")
                devengado_mes_series = _safe_numeric("devengado_mes")
                programado_mes_series = _safe_numeric("programado_mes")
                pivot_source_df = pd.DataFrame(
                    {
                        "sec_func": reporte_base["sec_func"].fillna("").astype(str),
                        "Generica": reporte_base["generica"].fillna("").astype(str),
                        "clasificador_cod-concepto": reporte_base["clasificador_cod_concepto"].fillna("").astype(str),
                    }
                )
                pivot_source_df["PIM"] = _safe_numeric("mto_pim")
                pivot_source_df["CERTIFICADO"] = _safe_numeric("mto_certificado")
                pivot_source_df["COMPROMETIDO"] = _safe_numeric("mto_compro_anual")
                pivot_source_df["DEVENGADO"] = devengado_acum
                pivot_source_df["DEVENGADO MES"] = devengado_mes_series
                pivot_source_df["PROGRAMADO MES"] = programado_mes_series
                pivot_source_df["NO CERTIFICADO"] = _safe_numeric("no_certificado")
                pivot_source_df["AVANCE"] = np.where(
                    pivot_source_df["PIM"] > 0,
                    devengado_acum / pivot_source_df["PIM"],
                    0.0,
                )
                pivot_source_df["AVANCE MES"] = np.where(
                    pivot_source_df["PROGRAMADO MES"] > 0,
                    devengado_mes_series / pivot_source_df["PROGRAMADO MES"],
                    0.0,
                )
                pivot_source_df = pivot_source_df[
                    [
                        "sec_func",
                        "Generica",
                        "clasificador_cod-concepto",
                        "PIM",
                        "CERTIFICADO",
                        "COMPROMETIDO",
                        "DEVENGADO",
                        "AVANCE",
                        "DEVENGADO MES",
                        "PROGRAMADO MES",
                        "AVANCE MES",
                        "NO CERTIFICADO",
                    ]
                ]
                reporte_siaf_pivot_source = pivot_source_df
            else:
                reporte_siaf_pivot_source = pd.DataFrame()

            def _label_or_default(value, fallback):
                text = "" if pd.isna(value) else str(value).strip()
                return text if text else fallback

            def _sort_key(label):
                text = _label_or_default(label, "")
                code = extract_code(text)
                if not code:
                    return (tuple(), text)
                parts = []
                for segment in code.split('.'):
                    try:
                        parts.append(int(segment))
                    except ValueError:
                        parts.append(segment)
                return (tuple(parts), text)

            def _format_label(level, text):
                indent = "    " * max(level, 0)
                prefix_map = {0: "", 1: "‚Ä¢ ", 2: "- "}
                return f"{indent}{prefix_map.get(level, '- ')}{text}".rstrip()

            records = []
            order_counter = 0

            for sec_value, sec_group in reporte_base.groupby("sec_func", sort=True):
                sec_label = _label_or_default(sec_value, "Sin sec_func")

                def _sum_metric(frame, source):
                    return float(frame[source].fillna(0.0).sum()) if source in frame.columns else 0.0

                sec_metrics = {dest: _sum_metric(sec_group, src) for dest, src in value_sources.items()}
                records.append(
                    {
                        "nivel": 0,
                        "orden": order_counter,
                        "Centro de costo / Gen√©rica de Gasto / Espec√≠fica de Gasto": _format_label(0, sec_label),
                        **sec_metrics,
                    }
                )
                order_counter += 1

                gen_groups = sorted(
                    sec_group.groupby("generica", dropna=False),
                    key=lambda kv: _sort_key(kv[0]),
                )

                for gen_value, gen_group in gen_groups:
                    gen_label = _label_or_default(gen_value, "Sin gen√©rica")
                    gen_metrics = {dest: _sum_metric(gen_group, src) for dest, src in value_sources.items()}
                    records.append(
                        {
                            "nivel": 1,
                            "orden": order_counter,
                            "Centro de costo / Gen√©rica de Gasto / Espec√≠fica de Gasto": _format_label(1, gen_label),
                            **gen_metrics,
                        }
                    )
                    order_counter += 1

                    detail_rows = sorted(
                        gen_group.to_dict("records"),
                        key=lambda row: _sort_key(row.get("especifica_det", "")),
                    )
                    for detail_row in detail_rows:
                        spec_label = _label_or_default(detail_row.get("clasificador_cod_concepto", ""), "Sin espec√≠fica")
                        if not spec_label or spec_label == "Sin espec√≠fica":
                            continue
                        detail_metrics = {
                            dest: float(detail_row.get(src, 0.0) or 0.0)
                            for dest, src in value_sources.items()
                        }
                        records.append(
                            {
                                "nivel": 2,
                                "orden": order_counter,
                                "Centro de costo / Gen√©rica de Gasto / Espec√≠fica de Gasto": _format_label(2, spec_label),
                                **detail_metrics,
                            }
                        )
                        order_counter += 1

            if records:
                reporte_siaf_df = pd.DataFrame.from_records(records)
                reporte_siaf_df["AVANCE"] = np.where(
                    reporte_siaf_df["PIM"].astype(float) > 0,
                    reporte_siaf_df["DEVENGADO"].astype(float) / reporte_siaf_df["PIM"].astype(float) * 100.0,
                    0.0,
                )
                reporte_siaf_df["AVANCE MES"] = np.where(
                    reporte_siaf_df["PROGRAMADO MES"].astype(float) > 0,
                    reporte_siaf_df["DEVENGADO MES"].astype(float)
                    / reporte_siaf_df["PROGRAMADO MES"].astype(float)
                    * 100.0,
                    0.0,
                )
                reporte_siaf_df = (
                    reporte_siaf_df.sort_values("orden", kind="stable")
                    .drop(columns=["orden", "nivel"], errors="ignore")
                )
                reporte_siaf_df = reporte_siaf_df[
                    [
                        "Centro de costo / Gen√©rica de Gasto / Espec√≠fica de Gasto",
                        "PIM",
                        "CERTIFICADO",
                        "COMPROMETIDO",
                        "DEVENGADO",
                        "AVANCE",
                        "DEVENGADO MES",
                        "PROGRAMADO MES",
                        "AVANCE MES",
                        "NO CERTIFICADO",
                    ]
                ]
            else:
                reporte_siaf_df = pd.DataFrame(
                    columns=[
                        "Centro de costo / Gen√©rica de Gasto / Espec√≠fica de Gasto",
                        "PIM",
                        "CERTIFICADO",
                        "COMPROMETIDO",
                        "DEVENGADO",
                        "AVANCE",
                        "DEVENGADO MES",
                        "PROGRAMADO MES",
                        "AVANCE MES",
                        "NO CERTIFICADO",
                    ]
                )

            reporte_display = standardize_financial_columns(reporte_siaf_df)
            reporte_display = round_numeric_for_reporting(reporte_display)
            fmt_reporte = build_style_formatters(reporte_display)
            reporte_style = reporte_display.style
            highlight_cols = [
                col
                for col in ["AVANCE", "AVANCE MES"]
                if col in reporte_display.columns
            ]
            if highlight_cols:
                reporte_style = reporte_style.applymap(
                    lambda v: "background-color: rgba(255, 41, 109, 0.35); color: #fff;"
                    if v < float(riesgo_umbral)
                    else "",
                    subset=highlight_cols,
                )
            if fmt_reporte:
                reporte_style = reporte_style.format(fmt_reporte)
            st.dataframe(reporte_style, use_container_width=True)
            total_detalles = int(reporte_base.shape[0])
            tot_pim = float(reporte_base.get("mto_pim", pd.Series(dtype=float)).sum()) if "mto_pim" in reporte_base else 0.0
            tot_dev = float(reporte_base.get("devengado", pd.Series(dtype=float)).sum()) if "devengado" in reporte_base else 0.0
            tot_cert = float(reporte_base.get("mto_certificado", pd.Series(dtype=float)).sum()) if "mto_certificado" in reporte_base else 0.0
            avance_total = (tot_dev / tot_pim * 100.0) if tot_pim else 0.0
            render_analysis(
                "El reporte detalla "
                f"{total_detalles} combinaciones de sec_func y gen√©rica; en conjunto acumulan S/ {tot_dev:,.2f} devengados "
                f"sobre un PIM de S/ {tot_pim:,.2f} (avance {avance_total:.2f}%), con S/ {tot_cert:,.2f} certificados.",
                icon="üßæ",
            )

with tab_descarga:
    with neon_panel(
        "Descarga de reportes",
        icon="‚¨áÔ∏è",
        subtitle="Exporta el tablero completo con el mismo brillo que la experiencia interactiva",
    ):
        if not XLSXWRITER_AVAILABLE:
            st.warning(
                "No se encontr√≥ la librer√≠a `xlsxwriter`. El Excel se generar√° sin tablas ni gr√°ficos embebidos."
            )
        excel_buffer = None
        excel_engine = None
        try:
            excel_buffer, excel_engine = to_excel_download(
                resumen=round_numeric_for_reporting(standardize_financial_columns(pivot.copy())),
                avance=round_numeric_for_reporting(avance_series.copy()),
                proyeccion=proyeccion_wide,
                ritmo=round_numeric_for_reporting(ritmo_df.copy()),
                leaderboard=round_numeric_for_reporting(standardize_financial_columns(leaderboard_df.copy())),
                reporte_siaf=round_numeric_for_reporting(standardize_financial_columns(reporte_siaf_df.copy())),
                reporte_siaf_pivot_source=reporte_siaf_pivot_source.copy(),
            )
        except ModuleNotFoundError as exc:
            missing = getattr(exc, "name", "xlsxwriter/openpyxl")
            st.error(
                "No se pudo generar el archivo de Excel porque faltan dependencias instaladas: "
                f"{missing}. Solicita al administrador que agregue el paquete correspondiente."
            )
        except Exception as exc:
            st.error(f"No se pudo generar el archivo de Excel: {exc}")
        else:
            if excel_engine == "openpyxl" and XLSXWRITER_AVAILABLE:
                st.info(
                    "`xlsxwriter` no se pudo inicializar, se utiliz√≥ `openpyxl` como alternativa. "
                    "Instala `xlsxwriter` para recuperar tablas y gr√°ficos embebidos."
                )

        if excel_buffer is not None:
            st.download_button(
                "Descargar Excel (Resumen + Avance)",
                data=excel_buffer,
                file_name="siaf_resumen_avance.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            if excel_engine:
                render_analysis(
                    "La descarga incluye todas las tablas mostradas en los apartados; "
                    f"se gener√≥ utilizando el motor de Excel `{excel_engine}`.",
                    icon="üíæ",
                )

